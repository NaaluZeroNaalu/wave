# import streamlit as st
# import subprocess
# import pandas as pd
# import os

# st.title("Wave Infra Application Launcher")

# # List of your two applications
# apps = {
#     "NCR App": "ncr.py",
#     "Report App": "appp.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to combine Excel outputs
# def combine_excel_outputs():
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#     ]
    
#     combined_df = pd.DataFrame()
#     for file in output_files:
#         if os.path.exists(file):
#             try:
#                 df = pd.read_excel(file)
#                 combined_df = pd.concat([combined_df, df], ignore_index=True, sort=False)
#                 st.write(f"Successfully read {file}")
#             except Exception as e:
#                 st.warning(f"Error reading {file}: {str(e)}. Skipping...")
#         else:
#             st.warning(f"{file} not found. Skipping...")

#     if not combined_df.empty:
#         combined_df.to_excel("final_output.xlsx", index=False)
#         st.success("Final Excel file generated as 'final_output.xlsx'!")
#         st.download_button(
#             label="Download Final Excel",
#             data=open("final_output.xlsx", "rb").read(),
#             file_name="final_output.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#     else:
#         st.error("No output files found to combine.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs, then click 'Combine All Outputs'.")


# import streamlit as st
# import subprocess
# import pandas as pd
# import os

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your two applications
# apps = {
#     "NCR App": "ncr.py",
#     "Report App": "appp.py",  # Corrected to Main.py based on earlier context
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to combine Excel outputs
# def combine_excel_outputs():
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#     ]
    
#     combined_df = pd.DataFrame()
#     for file in output_files:
#         if os.path.exists(file):
#             try:
#                 df = pd.read_excel(file)
#                 combined_df = pd.concat([combined_df, df], ignore_index=True, sort=False)
#                 st.write(f"Successfully read {file}")
#             except Exception as e:
#                 st.warning(f"Error reading {file}: {str(e)}. Skipping...")
#         else:
#             st.warning(f"{file} not found. Skipping...")

#     if not combined_df.empty:
#         combined_df.to_excel("final_output.xlsx", index=False)
#         st.success("Final Excel file generated as 'final_output.xlsx'!")
#         st.download_button(
#             label="Download Final Excel",
#             data=open("final_output.xlsx", "rb").read(),
#             file_name="final_output.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#     else:
#         st.error("No output files found to combine.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs, then click 'Combine All Outputs'.")


# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your two applications
# apps = {
#     "NCR App": "ncr.py",
#     "Report App": "Main.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to generate NCR Excel sheet
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR"):
#     output = io.BytesIO()
#     with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
#         workbook = writer.book
        
#         # Define formats
#         title_format = workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'fg_color': '#FFFF00',
#             'border': 1,
#             'font_size': 12
#         })
        
#         header_format = workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1,
#             'text_wrap': True
#         })
        
#         subheader_format = workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
        
#         cell_format = workbook.add_format({
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
        
#         site_format = workbook.add_format({
#             'align': 'left',
#             'valign': 'vcenter',
#             'border': 1
#         })
        
#         # Create worksheet
#         worksheet = workbook.add_worksheet('NCR Report')
        
#         # Set column widths
#         worksheet.set_column('A:A', 20)  # Site column
#         worksheet.set_column('B:I', 12)  # Data columns (8 data columns + Total)
        
#         # Get all unique sites from both sections
#         all_sites = set()
#         resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#         open_data = combined_result.get("NCR open beyond 21 days", {})
        
#         if not isinstance(resolved_data, dict) or "error" in resolved_data:
#             resolved_data = {"Sites": {}}
#         if not isinstance(open_data, dict) or "error" in open_data:
#             open_data = {"Sites": {}}
            
#         resolved_sites = resolved_data.get("Sites", {})
#         open_sites = open_data.get("Sites", {})
        
#         all_sites.update(resolved_sites.keys())
#         all_sites.update(open_sites.keys())
        
#         # Normalize site names
#         normalized_sites = {}
#         for site in all_sites:
#             match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#             if match:
#                 num = match.group(1)
#                 normalized_name = f"Veridia- Tower {num.zfill(2)}"
#                 normalized_sites[site] = normalized_name
#             else:
#                 normalized_sites[site] = site
                
#         # Add standard sites
#         standard_sites = [
#             "Veridia-Club",
#             "Veridia- Tower 01",
#             "Veridia- Tower 02",
#             "Veridia- Tower 03",
#             "Veridia- Tower 04 (A)",
#             "Veridia- Tower 04 (B)",
#             "Veridia- Tower 05",
#             "Veridia- Tower 06",
#             "Veridia- Tower 07",
#             "Veridia-Commercial",
#             "External Development"
#         ]
        
#         for site in standard_sites:
#             if site not in normalized_sites.values():
#                 normalized_sites[site] = site
        
#         # Sort the normalized sites
#         sorted_sites = sorted(list(set(normalized_sites.values())))
        
#         # Title row
#         month_match = re.search(r'NCR: (\w+)', report_title)
#         month = month_match.group(1) if month_match else datetime.now().strftime("%B")
#         worksheet.merge_range('A1:I1', f'NCR: {month}', title_format)
        
#         # Header row
#         row = 1
#         worksheet.write(row, 0, 'Site', header_format)
#         worksheet.merge_range(row, 1, row, 4, 'NCR resolved beyond 21 days', header_format)
#         worksheet.merge_range(row, 5, row, 7, 'NCR open beyond 21 days', header_format)
#         worksheet.write(row, 8, 'Total', header_format)
        
#         # Subheaders
#         row = 2
#         categories = ['Civil Finishing', 'MEP', 'Structure']
#         worksheet.write(row, 0, '', header_format)
        
#         # Resolved subheaders (4 columns: Civil Finishing, MEP, Structure, extra placeholder)
#         for i, cat in enumerate(categories + ['']):
#             worksheet.write(row, i + 1, cat, subheader_format)
        
#         # Open subheaders (3 columns: Civil Finishing, MEP, Structure)
#         for i, cat in enumerate(categories):
#             worksheet.write(row, i + 5, cat, subheader_format)
        
#         worksheet.write(row, 8, '', header_format)
        
#         # Map our categories to the JSON data categories
#         category_map = {
#             'Civil Finishing': 'FW',
#             'MEP': 'MEP',
#             'Structure': 'SW'
#         }
        
#         # Data rows
#         row = 3
#         site_totals = {}
        
#         for site in sorted_sites:
#             worksheet.write(row, 0, site, site_format)
            
#             # Find original site key
#             original_resolved_key = None
#             original_open_key = None
            
#             for orig_site, norm_site in normalized_sites.items():
#                 if norm_site == site:
#                     if orig_site in resolved_sites:
#                         original_resolved_key = orig_site
#                     if orig_site in open_sites:
#                         original_open_key = orig_site
            
#             site_total = 0
            
#             # Resolved data (4 columns: 3 categories + extra 0)
#             resolved_values = [0, 0, 0, 0]  # Initialize with 4 values
#             if original_resolved_key:
#                 for i, (display_cat, json_cat) in enumerate(category_map.items()):
#                     resolved_values[i] = resolved_sites.get(original_resolved_key, {}).get(json_cat, 0)
#             for i, value in enumerate(resolved_values):
#                 worksheet.write(row, i + 1, value, cell_format)
#                 site_total += value
                
#             # Open data (3 columns)
#             open_values = [0, 0, 0]
#             if original_open_key:
#                 for i, (display_cat, json_cat) in enumerate(category_map.items()):
#                     open_values[i] = open_sites.get(original_open_key, {}).get(json_cat, 0)
#             for i, value in enumerate(open_values):
#                 worksheet.write(row, i + 5, value, cell_format)
#                 site_total += value
                
#             # Total for this site
#             worksheet.write(row, 8, site_total, cell_format)
#             site_totals[site] = site_total
#             row += 1
        
#     output.seek(0)
#     return output

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#     ]
    
#     # Check if ncr_output.xlsx exists
#     if not os.path.exists(output_files[0]):
#         st.warning(f"{output_files[0]} not found. Skipping...")
#         return

#     # Read ncr_output.xlsx to reconstruct combined_result
#     try:
#         wb = openpyxl.load_workbook(output_files[0])
#         ws = wb['NCR Report']
        
#         # Extract data starting from row 4 (after headers)
#         data = []
#         for row in ws.iter_rows(min_row=4, values_only=True):
#             data.append(list(row))
        
#         # Reconstruct combined_result (simplified assumption based on structure)
#         combined_result = {
#             "NCR resolved beyond 21 days": {"Sites": {}},
#             "NCR open beyond 21 days": {"Sites": {}}
#         }
        
#         sites = [row[0] for row in data if row[0]]
#         for i, site in enumerate(sites):
#             resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW', ''], data[i][1:5]) if v is not None and v != 0}
#             open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][5:8]) if v is not None and v != 0}
#             if resolved_data:
#                 combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#             if open_data:
#                 combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
        
#         # Generate NCR sheet
#         ncr_output = generate_consolidated_ncr_excel(combined_result, report_title="NCR: April")
#         with open("temp_ncr.xlsx", "wb") as f:
#             f.write(ncr_output.read())
        
#     except Exception as e:
#         logger.error(f"Error processing ncr_output.xlsx: {str(e)}")
#         st.error(f"Error processing ncr_output.xlsx: {str(e)}. Skipping...")
#         return

#     # Create a new Excel writer for final output
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Copy NCR sheet
#     with open("temp_ncr.xlsx", "rb") as f:
#         ncr_wb = openpyxl.load_workbook(f)
#         ncr_ws = ncr_wb['NCR Report']
#         worksheet_ncr = workbook.add_worksheet('NCR Report')
#         for row in ncr_ws.rows:
#             for cell in row:
#                 worksheet_ncr.write(cell.row - 1, cell.column - 1, cell.value)

#     # Add main_output.xlsx to a second sheet
#     if os.path.exists(output_files[1]):
#         try:
#             main_df = pd.read_excel(output_files[1])
#             main_df.to_excel(output, sheet_name='Report Summary', index=False)

#             # Apply basic formatting to the Report Summary sheet
#             worksheet_main = output.sheets['Report Summary']
#             worksheet_main.set_column('A:A', 20)
#             worksheet_main.set_column('B:Z', 12)
#             for col_num, value in enumerate(main_df.columns):
#                 worksheet_main.write(0, col_num, value, header_format)
#             for row_num in range(len(main_df)):
#                 worksheet_main.write(row_num + 1, 0, str(main_df.iloc[row_num, 0]) if pd.notna(main_df.iloc[row_num, 0]) else '', site_format)
#                 for col_num in range(1, len(main_df.columns)):
#                     value = main_df.iloc[row_num, col_num]
#                     if pd.isna(value):
#                         worksheet_main.write(row_num + 1, col_num, '', cell_format)
#                     elif isinstance(value, (int, float)):
#                         worksheet_main.write_number(row_num + 1, col_num, value, cell_format)
#                     else:
#                         worksheet_main.write(row_num + 1, col_num, str(value), cell_format)
#         except Exception as e:
#             logger.error(f"Error processing main_output.xlsx: {str(e)}")
#             st.warning(f"Error processing main_output.xlsx: {str(e)}. Skipping...")
#     else:
#         st.warning(f"{output_files[1]} not found. Only NCR Report will be included.")

#     # Save and close the Excel file
#     output.close()
#     os.remove("temp_ncr.xlsx")  # Clean up temporary file
#     st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#     st.download_button(
#         label="Download Final Excel",
#         data=open("final_output.xlsx", "rb").read(),
#         file_name="final_output.xlsx",
#         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs, then click 'Combine All Outputs'.")


#Working Currenteley
# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Define formats at module level
# title_format = None
# header_format = None
# subheader_format = None
# cell_format = None
# site_format = None

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your two applications
# apps = {
#     "NCR App": "ncr.py",
#     "Report App": "appp.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to generate NCR Excel sheet
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None):
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output = io.BytesIO()
#     with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
#         # Use the provided workbook if available, otherwise create a new one
#         local_workbook = writer.book if workbook is None else workbook
        
#         # Initialize formats if not already set
#         if title_format is None:
#             title_format = local_workbook.add_format({
#                 'bold': True,
#                 'align': 'center',
#                 'valign': 'vcenter',
#                 'fg_color': '#FFFF00',
#                 'border': 1,
#                 'font_size': 12
#             })
        
#         if header_format is None:
#             header_format = local_workbook.add_format({
#                 'bold': True,
#                 'align': 'center',
#                 'valign': 'vcenter',
#                 'border': 1,
#                 'text_wrap': True
#             })
        
#         if subheader_format is None:
#             subheader_format = local_workbook.add_format({
#                 'bold': True,
#                 'align': 'center',
#                 'valign': 'vcenter',
#                 'border': 1
#             })
        
#         if cell_format is None:
#             cell_format = local_workbook.add_format({
#                 'align': 'center',
#                 'valign': 'vcenter',
#                 'border': 1
#             })
        
#         if site_format is None:
#             site_format = local_workbook.add_format({
#                 'align': 'left',
#                 'valign': 'vcenter',
#                 'border': 1
#             })
        
#         # Create worksheet
#         worksheet = local_workbook.add_worksheet('NCR Report') if workbook else writer.sheets['NCR Report']
        
#         # Set column widths
#         worksheet.set_column('A:A', 20)  # Site column
#         worksheet.set_column('B:I', 12)  # Data columns (8 data columns + Total)
        
#         # Get all unique sites from both sections
#         all_sites = set()
#         resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#         open_data = combined_result.get("NCR open beyond 21 days", {})
        
#         if not isinstance(resolved_data, dict) or "error" in resolved_data:
#             resolved_data = {"Sites": {}}
#         if not isinstance(open_data, dict) or "error" in open_data:
#             open_data = {"Sites": {}}
            
#         resolved_sites = resolved_data.get("Sites", {})
#         open_sites = open_data.get("Sites", {})
        
#         all_sites.update(resolved_sites.keys())
#         all_sites.update(open_sites.keys())
        
#         # Normalize site names
#         normalized_sites = {}
#         for site in all_sites:
#             match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#             if match:
#                 num = match.group(1)
#                 normalized_name = f"Veridia- Tower {num.zfill(2)}"
#                 normalized_sites[site] = normalized_name
#             else:
#                 normalized_sites[site] = site
                
#         # Add standard sites
#         standard_sites = [
#             "Veridia-Club",
#             "Veridia- Tower 01",
#             "Veridia- Tower 02",
#             "Veridia- Tower 03",
#             "Veridia- Tower 04 (A)",
#             "Veridia- Tower 04 (B)",
#             "Veridia- Tower 05",
#             "Veridia- Tower 06",
#             "Veridia- Tower 07",
#             "Veridia-Commercial",
#             "External Development"
#         ]
        
#         for site in standard_sites:
#             if site not in normalized_sites.values():
#                 normalized_sites[site] = site
        
#         # Sort the normalized sites
#         sorted_sites = sorted(list(set(normalized_sites.values())))
        
#         # Title row
#         month_match = re.search(r'NCR: (\w+)', report_title)
#         month = month_match.group(1) if month_match else datetime.now().strftime("%B")
#         worksheet.merge_range('A1:I1', f'NCR: {month}', title_format)
        
#         # Header row
#         row = 1
#         worksheet.write(row, 0, 'Site', header_format)
#         worksheet.merge_range(row, 1, row, 4, 'NCR resolved beyond 21 days', header_format)
#         worksheet.merge_range(row, 5, row, 7, 'NCR open beyond 21 days', header_format)
#         worksheet.write(row, 8, 'Total', header_format)
        
#         # Subheaders
#         row = 2
#         categories = ['Civil Finishing', 'MEP', 'Structure']
#         worksheet.write(row, 0, '', header_format)
        
#         # Resolved subheaders (4 columns: Civil Finishing, MEP, Structure, extra placeholder)
#         for i, cat in enumerate(categories + ['']):
#             worksheet.write(row, i + 1, cat, subheader_format)
        
#         # Open subheaders (3 columns: Civil Finishing, MEP, Structure)
#         for i, cat in enumerate(categories):
#             worksheet.write(row, i + 5, cat, subheader_format)
        
#         worksheet.write(row, 8, '', header_format)
        
#         # Map our categories to the JSON data categories
#         category_map = {
#             'Civil Finishing': 'FW',
#             'MEP': 'MEP',
#             'Structure': 'SW'
#         }
        
#         # Data rows
#         row = 3
#         site_totals = {}
        
#         for site in sorted_sites:
#             worksheet.write(row, 0, site, site_format)
            
#             # Find original site key
#             original_resolved_key = None
#             original_open_key = None
            
#             for orig_site, norm_site in normalized_sites.items():
#                 if norm_site == site:
#                     if orig_site in resolved_sites:
#                         original_resolved_key = orig_site
#                     if orig_site in open_sites:
#                         original_open_key = orig_site
            
#             site_total = 0
            
#             # Resolved data (4 columns: 3 categories + extra 0)
#             resolved_values = [0, 0, 0, 0]  # Initialize with 4 values
#             if original_resolved_key:
#                 for i, (display_cat, json_cat) in enumerate(category_map.items()):
#                     resolved_values[i] = resolved_sites.get(original_resolved_key, {}).get(json_cat, 0)
#             for i, value in enumerate(resolved_values):
#                 worksheet.write(row, i + 1, value, cell_format)
#                 site_total += value
                
#             # Open data (3 columns)
#             open_values = [0, 0, 0]
#             if original_open_key:
#                 for i, (display_cat, json_cat) in enumerate(category_map.items()):
#                     open_values[i] = open_sites.get(original_open_key, {}).get(json_cat, 0)
#             for i, value in enumerate(open_values):
#                 worksheet.write(row, i + 5, value, cell_format)
#                 site_total += value
                
#             # Total for this site
#             worksheet.write(row, 8, site_total, cell_format)
#             site_totals[site] = site_total
#             row += 1
        
#     if workbook is None:
#         output.seek(0)
#         return output
#     return None  # Return None if using existing workbook

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#     ]
    
#     # Check if ncr_output.xlsx exists
#     if not os.path.exists(output_files[0]):
#         st.warning(f"{output_files[0]} not found. Skipping...")
#         return

#     # Create a new Excel writer for final output
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Read ncr_output.xlsx to reconstruct combined_result
#     try:
#         wb = openpyxl.load_workbook(output_files[0])
#         ws = wb['NCR Report']
        
#         # Extract data starting from row 4 (after headers)
#         data = []
#         for row in ws.iter_rows(min_row=4, values_only=True):
#             data.append(list(row))
        
#         # Reconstruct combined_result (simplified assumption based on structure)
#         combined_result = {
#             "NCR resolved beyond 21 days": {"Sites": {}},
#             "NCR open beyond 21 days": {"Sites": {}}
#         }
        
#         sites = [row[0] for row in data if row[0]]
#         for i, site in enumerate(sites):
#             resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW', ''], data[i][1:5]) if v is not None and v != 0}
#             open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][5:8]) if v is not None and v != 0}
#             if resolved_data:
#                 combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#             if open_data:
#                 combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
        
#         # Generate NCR sheet directly into the final workbook
#         generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
        
#     except Exception as e:
#         logger.error(f"Error processing ncr_output.xlsx: {str(e)}")
#         st.error(f"Error processing ncr_output.xlsx: {str(e)}. Skipping...")
#         output.close()
#         return

#     # Add main_output.xlsx to a second sheet
#     if os.path.exists(output_files[1]):
#         try:
#             main_df = pd.read_excel(output_files[1])
#             main_df.to_excel(output, sheet_name='Report Summary', index=False)

#             # Apply basic formatting to the Report Summary sheet
#             worksheet_main = output.sheets['Report Summary']
#             worksheet_main.set_column('A:A', 20)
#             worksheet_main.set_column('B:Z', 12)
#             for col_num, value in enumerate(main_df.columns):
#                 worksheet_main.write(0, col_num, value, header_format)
#             for row_num in range(len(main_df)):
#                 worksheet_main.write(row_num + 1, 0, str(main_df.iloc[row_num, 0]) if pd.notna(main_df.iloc[row_num, 0]) else '', site_format)
#                 for col_num in range(1, len(main_df.columns)):
#                     value = main_df.iloc[row_num, col_num]
#                     if pd.isna(value):
#                         worksheet_main.write(row_num + 1, col_num, '', cell_format)
#                     elif isinstance(value, (int, float)):
#                         worksheet_main.write_number(row_num + 1, col_num, value, cell_format)
#                     else:
#                         worksheet_main.write(row_num + 1, col_num, str(value), cell_format)
#         except Exception as e:
#             logger.error(f"Error processing main_output.xlsx: {str(e)}")
#             st.warning(f"Error processing main_output.xlsx: {str(e)}. Skipping...")
#     else:
#         st.warning(f"{output_files[1]} not found. Only NCR Report will be included.")

#     # Save and close the Excel file
#     output.close()
#     st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#     # Verify file integrity and offer download
#     try:
#         with open("final_output.xlsx", "rb") as f:
#             file_content = f.read()
#         st.download_button(
#             label="Download Final Excel",
#             data=file_content,
#             file_name="final_output.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#     except Exception as e:
#         logger.error(f"Error reading final_output.xlsx for download: {str(e)}")
#         st.error("Failed to generate a valid Excel file. Please check the logs.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs, then click 'Combine All Outputs'.")

#working correctely
# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Define formats at module level
# title_format = None
# header_format = None
# subheader_format = None
# cell_format = None
# site_format = None

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your two applications
# apps = {
#     "NCR App": "ncr.py",
#     "Report App": "appp.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to generate NCR Excel sheet (adapted from new ncr.py)
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None, worksheet=None):
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     # Use existing workbook and worksheet if provided, otherwise create new ones
#     if workbook is None:
#         output = io.BytesIO()
#         writer = pd.ExcelWriter(output, engine='xlsxwriter')
#         local_workbook = writer.book
#         local_worksheet = local_workbook.add_worksheet('NCR Report')
#     else:
#         local_workbook = workbook
#         local_worksheet = worksheet or local_workbook.add_worksheet('NCR Report')

#     # Initialize formats if not already set
#     if title_format is None:
#         title_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'fg_color': 'yellow',
#             'border': 1,
#             'font_size': 12
#         })
    
#     if header_format is None:
#         header_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1,
#             'text_wrap': True
#         })
    
#     if subheader_format is None:
#         subheader_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if cell_format is None:
#         cell_format = local_workbook.add_format({
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if site_format is None:
#         site_format = local_workbook.add_format({
#             'align': 'left',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     # Set column widths
#     local_worksheet.set_column('A:A', 20)  # Site column
#     local_worksheet.set_column('B:H', 12)  # Data columns (7 columns total)
    
#     # Get data from both sections
#     resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#     open_data = combined_result.get("NCR open beyond 21 days", {})
    
#     if not isinstance(resolved_data, dict) or "error" in resolved_data:
#         resolved_data = {"Sites": {}}
#     if not isinstance(open_data, dict) or "error" in open_data:
#         open_data = {"Sites": {}}
        
#     resolved_sites = resolved_data.get("Sites", {})
#     open_sites = open_data.get("Sites", {})
    
#     # Define only the standard sites you want to include
#     standard_sites = [
#         "Veridia-Club",
#         "Veridia- Tower 01",
#         "Veridia- Tower 02",
#         "Veridia- Tower 03",
#         "Veridia- Tower 04",
#         "Veridia- Tower 05",
#         "Veridia- Tower 06",
#         "Veridia- Tower 07",
#         "Veridia-Commercial",
#         "External Development"
#     ]
    
#     # Normalize JSON site names to match standard_sites format
#     def normalize_site_name(site):
#         if site in ["Veridia-Club", "Veridia-Commercial"]:
#             return site
#         match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#         if match:
#             num = match.group(1).zfill(2)
#             return f"Veridia- Tower {num}"
#         return site

#     # Create a reverse mapping for original keys to normalized names
#     site_mapping = {k: normalize_site_name(k) for k in (resolved_sites.keys() | open_sites.keys())}
    
#     # Sort the standard sites
#     sorted_sites = sorted(standard_sites)
    
#     # Title row
#     local_worksheet.merge_range('A1:H1', report_title, title_format)
    
#     # Header row
#     row = 1
#     local_worksheet.write(row, 0, 'Site', header_format)
#     local_worksheet.merge_range(row, 1, row, 3, 'NCR resolved beyond 21 days', header_format)
#     local_worksheet.merge_range(row, 4, row, 6, 'NCR open beyond 21 days', header_format)
#     local_worksheet.write(row, 7, 'Total', header_format)
    
#     # Subheaders
#     row = 2
#     categories = ['Civil Finishing', 'MEP', 'Structure']
#     local_worksheet.write(row, 0, '', header_format)
    
#     # Resolved subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+1, cat, subheader_format)
        
#     # Open subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+4, cat, subheader_format)
        
#     local_worksheet.write(row, 7, '', header_format)
    
#     # Map our categories to the JSON data categories
#     category_map = {
#         'Civil Finishing': 'FW',
#         'MEP': 'MEP',
#         'Structure': 'SW'
#     }
    
#     # Data rows
#     row = 3
#     site_totals = {}
    
#     for site in sorted_sites:
#         local_worksheet.write(row, 0, site, site_format)
        
#         # Find original key that maps to this normalized site
#         original_resolved_key = next((k for k, v in site_mapping.items() if v == site), None)
#         original_open_key = next((k for k, v in site_mapping.items() if v == site), None)
        
#         site_total = 0
        
#         # Resolved data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_resolved_key and original_resolved_key in resolved_sites:
#                 value = resolved_sites[original_resolved_key].get(json_cat, 0)
#             local_worksheet.write(row, i+1, value, cell_format)
#             site_total += value
            
#         # Open data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_open_key and original_open_key in open_sites:
#                 value = open_sites[original_open_key].get(json_cat, 0)
#             local_worksheet.write(row, i+4, value, cell_format)
#             site_total += value
            
#         # Total for this site
#         local_worksheet.write(row, 7, site_total, cell_format)
#         site_totals[site] = site_total
#         row += 1
    
#     if workbook is None:
#         writer.save()
#         output.seek(0)
#         return output
#     return None  # No return needed when using existing workbook

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#     ]
    
#     # Check if ncr_output.xlsx exists
#     if not os.path.exists(output_files[0]):
#         st.warning(f"{output_files[0]} not found. Skipping...")
#         return

#     # Create a new Excel writer for final output
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Read ncr_output.xlsx to reconstruct combined_result
#     try:
#         wb = openpyxl.load_workbook(output_files[0])
#         ws = wb['NCR Report']
        
#         # Extract data starting from row 4 (after headers)
#         data = []
#         for row in ws.iter_rows(min_row=4, values_only=True):
#             data.append(list(row))
        
#         # Reconstruct combined_result (simplified assumption based on structure)
#         combined_result = {
#             "NCR resolved beyond 21 days": {"Sites": {}},
#             "NCR open beyond 21 days": {"Sites": {}}
#         }
        
#         sites = [row[0] for row in data if row[0]]
#         for i, site in enumerate(sites):
#             resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][1:4]) if v is not None and v != 0}
#             open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][4:7]) if v is not None and v != 0}
#             if resolved_data:
#                 combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#             if open_data:
#                 combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
        
#         # Generate NCR sheet directly into the final workbook
#         generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
        
#     except Exception as e:
#         logger.error(f"Error processing ncr_output.xlsx: {str(e)}")
#         st.error(f"Error processing ncr_output.xlsx: {str(e)}. Skipping...")
#         output.close()
#         return

#     # Add main_output.xlsx to a second sheet
#     if os.path.exists(output_files[1]):
#         try:
#             main_df = pd.read_excel(output_files[1])
#             main_df.to_excel(output, sheet_name='Report Summary', index=False, startrow=0)

#             # Apply basic formatting to the Report Summary sheet
#             worksheet_main = output.sheets['Report Summary']
#             worksheet_main.set_column('A:A', 20)
#             worksheet_main.set_column('B:Z', 12)
#             for col_num, value in enumerate(main_df.columns):
#                 worksheet_main.write(0, col_num, value, header_format)
#             for row_num in range(len(main_df)):
#                 worksheet_main.write(row_num, 0, str(main_df.iloc[row_num, 0]) if pd.notna(main_df.iloc[row_num, 0]) else '', site_format)
#                 for col_num in range(1, len(main_df.columns)):
#                     value = main_df.iloc[row_num, col_num]
#                     if pd.isna(value):
#                         worksheet_main.write(row_num, col_num, '', cell_format)
#                     elif isinstance(value, (int, float)):
#                         worksheet_main.write_number(row_num, col_num, value, cell_format)
#                     else:
#                         worksheet_main.write(row_num, col_num, str(value), cell_format)
#         except Exception as e:
#             logger.error(f"Error processing main_output.xlsx: {str(e)}")
#             st.warning(f"Error processing main_output.xlsx: {str(e)}. Skipping...")
#     else:
#         st.warning(f"{output_files[1]} not found. Only NCR Report will be included.")

#     # Save and close the Excel file
#     output.close()
#     st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#     # Verify file integrity and offer download
#     try:
#         with open("final_output.xlsx", "rb") as f:
#             file_content = f.read()
#         st.download_button(
#             label="Download Final Excel",
#             data=file_content,
#             file_name="final_output.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#     except Exception as e:
#         logger.error(f"Error reading final_output.xlsx for download: {str(e)}")
#         st.error("Failed to generate a valid Excel file. Please check the logs.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs, then click 'Combine All Outputs'.")

#Working Correcteley
# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Define formats at module level
# title_format = None
# header_format = None
# subheader_format = None
# cell_format = None
# site_format = None

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your three applications
# apps = {
#     "NCR App": "ncr.py",
#     "Checklist App": "Checklist.py",
#     "Report App": "appp.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to generate NCR Excel sheet (adapted from new ncr.py)
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None, worksheet=None):
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     # Use existing workbook and worksheet if provided, otherwise create new ones
#     if workbook is None:
#         output = io.BytesIO()
#         writer = pd.ExcelWriter(output, engine='xlsxwriter')
#         local_workbook = writer.book
#         local_worksheet = local_workbook.add_worksheet('NCR Report')
#     else:
#         local_workbook = workbook
#         local_worksheet = worksheet or local_workbook.add_worksheet('NCR Report')

#     # Initialize formats if not already set
#     if title_format is None:
#         title_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'fg_color': 'yellow',
#             'border': 1,
#             'font_size': 12
#         })
    
#     if header_format is None:
#         header_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1,
#             'text_wrap': True
#         })
    
#     if subheader_format is None:
#         subheader_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if cell_format is None:
#         cell_format = local_workbook.add_format({
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if site_format is None:
#         site_format = local_workbook.add_format({
#             'align': 'left',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     # Set column widths
#     local_worksheet.set_column('A:A', 20)  # Site column
#     local_worksheet.set_column('B:H', 12)  # Data columns (7 columns total)
    
#     # Get data from both sections
#     resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#     open_data = combined_result.get("NCR open beyond 21 days", {})
    
#     if not isinstance(resolved_data, dict) or "error" in resolved_data:
#         resolved_data = {"Sites": {}}
#     if not isinstance(open_data, dict) or "error" in open_data:
#         open_data = {"Sites": {}}
        
#     resolved_sites = resolved_data.get("Sites", {})
#     open_sites = open_data.get("Sites", {})
    
#     # Define only the standard sites you want to include
#     standard_sites = [
#         "Veridia-Club",
#         "Veridia- Tower 01",
#         "Veridia- Tower 02",
#         "Veridia- Tower 03",
#         "Veridia- Tower 04",
#         "Veridia- Tower 05",
#         "Veridia- Tower 06",
#         "Veridia- Tower 07",
#         "Veridia-Commercial",
#         "External Development"
#     ]
    
#     # Normalize JSON site names to match standard_sites format
#     def normalize_site_name(site):
#         if site in ["Veridia-Club", "Veridia-Commercial"]:
#             return site
#         match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#         if match:
#             num = match.group(1).zfill(2)
#             return f"Veridia- Tower {num}"
#         return site

#     # Create a reverse mapping for original keys to normalized names
#     site_mapping = {k: normalize_site_name(k) for k in (resolved_sites.keys() | open_sites.keys())}
    
#     # Sort the standard sites
#     sorted_sites = sorted(standard_sites)
    
#     # Title row
#     local_worksheet.merge_range('A1:H1', report_title, title_format)
    
#     # Header row
#     row = 1
#     local_worksheet.write(row, 0, 'Site', header_format)
#     local_worksheet.merge_range(row, 1, row, 3, 'NCR resolved beyond 21 days', header_format)
#     local_worksheet.merge_range(row, 4, row, 6, 'NCR open beyond 21 days', header_format)
#     local_worksheet.write(row, 7, 'Total', header_format)
    
#     # Subheaders
#     row = 2
#     categories = ['Civil Finishing', 'MEP', 'Structure']
#     local_worksheet.write(row, 0, '', header_format)
    
#     # Resolved subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+1, cat, subheader_format)
        
#     # Open subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+4, cat, subheader_format)
        
#     local_worksheet.write(row, 7, '', header_format)
    
#     # Map our categories to the JSON data categories
#     category_map = {
#         'Civil Finishing': 'FW',
#         'MEP': 'MEP',
#         'Structure': 'SW'
#     }
    
#     # Data rows
#     row = 3
#     site_totals = {}
    
#     for site in sorted_sites:
#         local_worksheet.write(row, 0, site, site_format)
        
#         # Find original key that maps to this normalized site
#         original_resolved_key = next((k for k, v in site_mapping.items() if v == site), None)
#         original_open_key = next((k for k, v in site_mapping.items() if v == site), None)
        
#         site_total = 0
        
#         # Resolved data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_resolved_key and original_resolved_key in resolved_sites:
#                 value = resolved_sites[original_resolved_key].get(json_cat, 0)
#             local_worksheet.write(row, i+1, value, cell_format)
#             site_total += value
            
#         # Open data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_open_key and original_open_key in open_sites:
#                 value = open_sites[original_open_key].get(json_cat, 0)
#             local_worksheet.write(row, i+4, value, cell_format)
#             site_total += value
            
#         # Total for this site
#         local_worksheet.write(row, 7, site_total, cell_format)
#         site_totals[site] = site_total
#         row += 1
    
#     if workbook is None:
#         writer.save()
#         output.seek(0)
#         return output
#     return None  # No return needed when using existing workbook

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#     ]
    
#     # Check if ncr_output.xlsx exists
#     if not os.path.exists(output_files[0]):
#         st.warning(f"{output_files[0]} not found. Skipping...")
#         return

#     # Create a new Excel writer for final output
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Read ncr_output.xlsx to reconstruct combined_result
#     try:
#         wb = openpyxl.load_workbook(output_files[0])
#         ws = wb['NCR Report']
        
#         # Extract data starting from row 4 (after headers)
#         data = []
#         for row in ws.iter_rows(min_row=4, values_only=True):
#             data.append(list(row))
        
#         # Reconstruct combined_result (simplified assumption based on structure)
#         combined_result = {
#             "NCR resolved beyond 21 days": {"Sites": {}},
#             "NCR open beyond 21 days": {"Sites": {}}
#         }
        
#         sites = [row[0] for row in data if row[0]]
#         for i, site in enumerate(sites):
#             resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][1:4]) if v is not None and v != 0}
#             open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][4:7]) if v is not None and v != 0}
#             if resolved_data:
#                 combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#             if open_data:
#                 combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
        
#         # Generate NCR sheet directly into the final workbook
#         generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
        
#     except Exception as e:
#         logger.error(f"Error processing ncr_output.xlsx: {str(e)}")
#         st.error(f"Error processing ncr_output.xlsx: {str(e)}. Skipping...")
#         output.close()
#         return

#     # Add main_output.xlsx to a second sheet
#     if os.path.exists(output_files[1]):
#         try:
#             main_df = pd.read_excel(output_files[1])
#             main_df.to_excel(output, sheet_name='Report Summary', index=False, startrow=0)

#             # Apply basic formatting to the Report Summary sheet
#             worksheet_main = output.sheets['Report Summary']
#             worksheet_main.set_column('A:A', 20)
#             worksheet_main.set_column('B:Z', 12)
#             for col_num, value in enumerate(main_df.columns):
#                 worksheet_main.write(0, col_num, value, header_format)
#             for row_num in range(len(main_df)):
#                 worksheet_main.write(row_num, 0, str(main_df.iloc[row_num, 0]) if pd.notna(main_df.iloc[row_num, 0]) else '', site_format)
#                 for col_num in range(1, len(main_df.columns)):
#                     value = main_df.iloc[row_num, col_num]
#                     if pd.isna(value):
#                         worksheet_main.write(row_num, col_num, '', cell_format)
#                     elif isinstance(value, (int, float)):
#                         worksheet_main.write_number(row_num, col_num, value, cell_format)
#                     else:
#                         worksheet_main.write(row_num, col_num, str(value), cell_format)
#         except Exception as e:
#             logger.error(f"Error processing main_output.xlsx: {str(e)}")
#             st.warning(f"Error processing main_output.xlsx: {str(e)}. Skipping...")
#     else:
#         st.warning(f"{output_files[1]} not found. Only NCR Report will be included.")

#     # Save and close the Excel file
#     output.close()
#     st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#     # Verify file integrity and offer download
#     try:
#         with open("final_output.xlsx", "rb") as f:
#             file_content = f.read()
#         st.download_button(
#             label="Download Final Excel",
#             data=file_content,
#             file_name="final_output.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#     except Exception as e:
#         logger.error(f"Error reading final_output.xlsx for download: {str(e)}")
#         st.error("Failed to generate a valid Excel file. Please check the logs.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs, then click 'Combine All Outputs'.")


# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime
# from project_main import final_output

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Define formats at module level
# title_format = None
# header_format = None
# subheader_format = None
# cell_format = None
# site_format = None

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your three applications
# apps = {
#     "NCR App": "ncr.py",
#     "Checklist App": "CheckList.py",
#     "Report App": "appp.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to generate NCR Excel sheet
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None, worksheet=None):
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     # Use existing workbook and worksheet if provided, otherwise create new ones
#     if workbook is None:
#         output = io.BytesIO()
#         writer = pd.ExcelWriter(output, engine='xlsxwriter')
#         local_workbook = writer.book
#         local_worksheet = local_workbook.add_worksheet('NCR Report')
#     else:
#         local_workbook = workbook
#         local_worksheet = worksheet or local_workbook.add_worksheet('NCR Report')

#     # Initialize formats if not already set
#     if title_format is None:
#         title_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'fg_color': 'yellow',
#             'border': 1,
#             'font_size': 12
#         })
    
#     if header_format is None:
#         header_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1,
#             'text_wrap': True
#         })
    
#     if subheader_format is None:
#         subheader_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if cell_format is None:
#         cell_format = local_workbook.add_format({
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if site_format is None:
#         site_format = local_workbook.add_format({
#             'align': 'left',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     # Set column widths
#     local_worksheet.set_column('A:A', 20)  # Site column
#     local_worksheet.set_column('B:H', 12)  # Data columns (7 columns total)
    
#     # Get data from both sections
#     resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#     open_data = combined_result.get("NCR open beyond 21 days", {})
    
#     if not isinstance(resolved_data, dict) or "error" in resolved_data:
#         resolved_data = {"Sites": {}}
#     if not isinstance(open_data, dict) or "error" in open_data:
#         open_data = {"Sites": {}}
        
#     resolved_sites = resolved_data.get("Sites", {})
#     open_sites = open_data.get("Sites", {})
    
#     # Define only the standard sites you want to include
#     standard_sites = [
#         "Veridia-Club",
#         "Veridia- Tower 01",
#         "Veridia- Tower 02",
#         "Veridia- Tower 03",
#         "Veridia- Tower 04",
#         "Veridia- Tower 05",
#         "Veridia- Tower 06",
#         "Veridia- Tower 07",
#         "Veridia-Commercial",
#         "External Development"
#     ]
    
#     # Normalize JSON site names to match standard_sites format
#     def normalize_site_name(site):
#         if site in ["Veridia-Club", "Veridia-Commercial"]:
#             return site
#         match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#         if match:
#             num = match.group(1).zfill(2)
#             return f"Veridia- Tower {num}"
#         return site

#     # Create a reverse mapping for original keys to normalized names
#     site_mapping = {k: normalize_site_name(k) for k in (resolved_sites.keys() | open_sites.keys())}
    
#     # Sort the standard sites
#     sorted_sites = sorted(standard_sites)
    
#     # Title row
#     local_worksheet.merge_range('A1:H1', report_title, title_format)
    
#     # Header row
#     row = 1
#     local_worksheet.write(row, 0, 'Site', header_format)
#     local_worksheet.merge_range(row, 1, row, 3, 'NCR resolved beyond 21 days', header_format)
#     local_worksheet.merge_range(row, 4, row, 6, 'NCR open beyond 21 days', header_format)
#     local_worksheet.write(row, 7, 'Total', header_format)
    
#     # Subheaders
#     row = 2
#     categories = ['Civil Finishing', 'MEP', 'Structure']
#     local_worksheet.write(row, 0, '', header_format)
    
#     # Resolved subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+1, cat, subheader_format)
        
#     # Open subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+4, cat, subheader_format)
        
#     local_worksheet.write(row, 7, '', header_format)
    
#     # Map our categories to the JSON data categories
#     category_map = {
#         'Civil Finishing': 'FW',
#         'MEP': 'MEP',
#         'Structure': 'SW'
#     }
    
#     # Data rows
#     row = 3
#     site_totals = {}
    
#     for site in sorted_sites:
#         local_worksheet.write(row, 0, site, site_format)
        
#         # Find original key that maps to this normalized site
#         original_resolved_key = next((k for k, v in site_mapping.items() if v == site), None)
#         original_open_key = next((k for k, v in site_mapping.items() if v == site), None)
        
#         site_total = 0
        
#         # Resolved data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_resolved_key and original_resolved_key in resolved_sites:
#                 value = resolved_sites[original_resolved_key].get(json_cat, 0)
#             local_worksheet.write(row, i+1, value, cell_format)
#             site_total += value
            
#         # Open data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_open_key and original_open_key in open_sites:
#                 value = open_sites[original_open_key].get(json_cat, 0)
#             local_worksheet.write(row, i+4, value, cell_format)
#             site_total += value
            
#         # Total for this site
#         local_worksheet.write(row, 7, site_total, cell_format)
#         site_totals[site] = site_total
#         row += 1
    
#     if workbook is None:
#         writer.save()
#         output.seek(0)
#         return output
#     return None  # No return needed when using existing workbook

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#     ]
    
#     # Check if ncr_output.xlsx exists
#     if not os.path.exists(output_files[0]):
#         logger.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")
#         st.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")
#         return

#     # Create a new Excel writer for final output
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Read ncr_output.xlsx to reconstruct combined_result
#     try:
#         ncr_wb = openpyxl.load_workbook(output_files[0])
#         ncr_ws = ncr_wb['NCR Report']
        
#         # Extract data starting from row 4 (after headers)
#         data = []
#         for row in ncr_ws.iter_rows(min_row=4, values_only=True):
#             data.append(list(row))
        
#         # Reconstruct combined_result (simplified assumption based on structure)
#         combined_result = {
#             "NCR resolved beyond 21 days": {"Sites": {}},
#             "NCR open beyond 21 days": {"Sites": {}}
#         }
        
#         sites = [row[0] for row in data if row[0]]
#         for i, site in enumerate(sites):
#             resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][1:4]) if v is not None and v != 0}
#             open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][4:7]) if v is not None and v != 0}
#             if resolved_data:
#                 combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#             if open_data:
#                 combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
        
#         # Generate NCR sheet directly into the final workbook
#         generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
#         logger.info("Successfully added NCR Report sheet to final_output.xlsx")
        
#     except Exception as e:
#         logger.error(f"Error processing ncr_output.xlsx: {str(e)}")
#         st.error(f"Error processing ncr_output.xlsx: {str(e)}. Skipping NCR Report...")
#         output.close()
#         return

#     # Add main_output.xlsx data
#     if os.path.exists(output_files[1]):
#         try:
#             main_wb = openpyxl.load_workbook(output_files[1])
#             # Copy all available sheets from main_output.xlsx
#             for sheet_name in ['Raw Data', 'Checklist Summary']:
#                 if sheet_name in main_wb.sheetnames:
#                     ws = main_wb[sheet_name]
#                     new_ws = workbook.add_worksheet(sheet_name)
#                     for row in ws.rows:
#                         for cell in row:
#                             new_ws.write(cell.row - 1, cell.column - 1, cell.value)
#                     logger.info(f"Successfully added {sheet_name} sheet from main_output.xlsx")
#                 else:
#                     logger.warning(f"Sheet '{sheet_name}' not found in main_output.xlsx. Skipping...")
#                     st.warning(f"Sheet '{sheet_name}' not found in main_output.xlsx. Skipping...")
#         except Exception as e:
#             logger.error(f"Error processing main_output.xlsx: {str(e)}")
#             st.warning(f"Error processing main_output.xlsx: {str(e)}. Skipping remaining sheets...")
#     else:
#         logger.warning(f"{output_files[1]} not found. Only NCR Report will be included.")
#         st.warning(f"{output_files[1]} not found. Only NCR Report will be included.")

#     # Save and close the Excel file
#     try:
#         output.save()
#         logger.info("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#         st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
        
#         # Verify file integrity and offer download
#         with open("final_output.xlsx", "rb") as f:
#             file_content = f.read()
#         st.download_button(
#             label="Download Final Excel",
#             data=file_content,
#             file_name="final_output.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#     except Exception as e:
#         logger.error(f"Error saving or reading final_output.xlsx for download: {str(e)}")
#         st.error("Failed to generate a valid Excel file. Please check the logs.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs (e.g., ncr_output.xlsx and main_output.xlsx), then click 'Combine All Outputs'.")



# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Define formats at module level
# title_format = None
# header_format = None
# subheader_format = None
# cell_format = None
# site_format = None

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your applications (updated with WatsonX Processor)
# apps = {
#     "NCR App": "ncr.py",
#     "Checklist App": "CheckList.py",
#     "Report App": "appp.py",
#     "WatsonX Processor": "project_main.py",  # New app for your provided code
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])  # Opens app in a new process

# # Function to generate NCR Excel sheet
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None, worksheet=None):
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     # Use existing workbook and worksheet if provided, otherwise create new ones
#     if workbook is None:
#         output = io.BytesIO()
#         writer = pd.ExcelWriter(output, engine='xlsxwriter')
#         local_workbook = writer.book
#         local_worksheet = local_workbook.add_worksheet('NCR Report')
#     else:
#         local_workbook = workbook
#         local_worksheet = worksheet or local_workbook.add_worksheet('NCR Report')

#     # Initialize formats if not already set
#     if title_format is None:
#         title_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'fg_color': 'yellow',
#             'border': 1,
#             'font_size': 12
#         })
    
#     if header_format is None:
#         header_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1,
#             'text_wrap': True
#         })
    
#     if subheader_format is None:
#         subheader_format = local_workbook.add_format({
#             'bold': True,
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if cell_format is None:
#         cell_format = local_workbook.add_format({
#             'align': 'center',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     if site_format is None:
#         site_format = local_workbook.add_format({
#             'align': 'left',
#             'valign': 'vcenter',
#             'border': 1
#         })
    
#     # Set column widths
#     local_worksheet.set_column('A:A', 20)  # Site column
#     local_worksheet.set_column('B:H', 12)  # Data columns (7 columns total)
    
#     # Get data from both sections
#     resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#     open_data = combined_result.get("NCR open beyond 21 days", {})
    
#     if not isinstance(resolved_data, dict) or "error" in resolved_data:
#         resolved_data = {"Sites": {}}
#     if not isinstance(open_data, dict) or "error" in open_data:
#         open_data = {"Sites": {}}
        
#     resolved_sites = resolved_data.get("Sites", {})
#     open_sites = open_data.get("Sites", {})
    
#     # Define only the standard sites you want to include
#     standard_sites = [
#         "Veridia-Club",
#         "Veridia- Tower 01",
#         "Veridia- Tower 02",
#         "Veridia- Tower 03",
#         "Veridia- Tower 04",
#         "Veridia- Tower 05",
#         "Veridia- Tower 06",
#         "Veridia- Tower 07",
#         "Veridia-Commercial",
#         "External Development"
#     ]
    
#     # Normalize JSON site names to match standard_sites format
#     def normalize_site_name(site):
#         if site in ["Veridia-Club", "Veridia-Commercial"]:
#             return site
#         match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#         if match:
#             num = match.group(1).zfill(2)
#             return f"Veridia- Tower {num}"
#         return site

#     # Create a reverse mapping for original keys to normalized names
#     site_mapping = {k: normalize_site_name(k) for k in (resolved_sites.keys() | open_sites.keys())}
    
#     # Sort the standard sites
#     sorted_sites = sorted(standard_sites)
    
#     # Title row
#     local_worksheet.merge_range('A1:H1', report_title, title_format)
    
#     # Header row
#     row = 1
#     local_worksheet.write(row, 0, 'Site', header_format)
#     local_worksheet.merge_range(row, 1, row, 3, 'NCR resolved beyond 21 days', header_format)
#     local_worksheet.merge_range(row, 4, row, 6, 'NCR open beyond 21 days', header_format)
#     local_worksheet.write(row, 7, 'Total', header_format)
    
#     # Subheaders
#     row = 2
#     categories = ['Civil Finishing', 'MEP', 'Structure']
#     local_worksheet.write(row, 0, '', header_format)
    
#     # Resolved subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+1, cat, subheader_format)
        
#     # Open subheaders
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+4, cat, subheader_format)
        
#     local_worksheet.write(row, 7, '', header_format)
    
#     # Map our categories to the JSON data categories
#     category_map = {
#         'Civil Finishing': 'FW',
#         'MEP': 'MEP',
#         'Structure': 'SW'
#     }
    
#     # Data rows
#     row = 3
#     site_totals = {}
    
#     for site in sorted_sites:
#         local_worksheet.write(row, 0, site, site_format)
        
#         # Find original key that maps to this normalized site
#         original_resolved_key = next((k for k, v in site_mapping.items() if v == site), None)
#         original_open_key = next((k for k, v in site_mapping.items() if v == site), None)
        
#         site_total = 0
        
#         # Resolved data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_resolved_key and original_resolved_key in resolved_sites:
#                 value = resolved_sites[original_resolved_key].get(json_cat, 0)
#             local_worksheet.write(row, i+1, value, cell_format)
#             site_total += value
            
#         # Open data
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_open_key and original_open_key in open_sites:
#                 value = open_sites[original_open_key].get(json_cat, 0)
#             local_worksheet.write(row, i+4, value, cell_format)
#             site_total += value
            
#         # Total for this site
#         local_worksheet.write(row, 7, site_total, cell_format)
#         site_totals[site] = site_total
#         row += 1
    
#     if workbook is None:
#         writer.save()
#         output.seek(0)
#         return output
#     return None  # No return needed when using existing workbook

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#         "project_main.xlsx",  # Wildcard for dynamic timestamp
#     ]
    
#     # Create a new Excel writer for final output
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Read ncr_output.xlsx to reconstruct combined_result
#     try:
#         if os.path.exists(output_files[0]):
#             ncr_wb = openpyxl.load_workbook(output_files[0])
#             ncr_ws = ncr_wb['NCR Report']
            
#             # Extract data starting from row 4 (after headers)
#             data = []
#             for row in ncr_ws.iter_rows(min_row=4, values_only=True):
#                 data.append(list(row))
            
#             # Reconstruct combined_result (simplified assumption based on structure)
#             combined_result = {
#                 "NCR resolved beyond 21 days": {"Sites": {}},
#                 "NCR open beyond 21 days": {"Sites": {}}
#             }
            
#             sites = [row[0] for row in data if row[0]]
#             for i, site in enumerate(sites):
#                 resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][1:4]) if v is not None and v != 0}
#                 open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][4:7]) if v is not None and v != 0}
#                 if resolved_data:
#                     combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#                 if open_data:
#                     combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
            
#             # Generate NCR sheet directly into the final workbook
#             generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
#             logger.info("Successfully added NCR Report sheet to final_output.xlsx")
#         else:
#             logger.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")
#             st.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")
#     except Exception as e:
#         logger.error(f"Error processing ncr_output.xlsx: {str(e)}")
#         st.error(f"Error processing ncr_output.xlsx: {str(e)}. Skipping NCR Report...")
#         output.close()
#         return

#     # Add main_output.xlsx data
#     if os.path.exists(output_files[1]):
#         try:
#             main_wb = openpyxl.load_workbook(output_files[1])
#             # Copy all available sheets from main_output.xlsx
#             for sheet_name in ['Raw Data', 'Checklist Summary']:
#                 if sheet_name in main_wb.sheetnames:
#                     ws = main_wb[sheet_name]
#                     new_ws = workbook.add_worksheet(sheet_name)
#                     for row in ws.rows:
#                         for cell in row:
#                             new_ws.write(cell.row - 1, cell.column - 1, cell.value)
#                     logger.info(f"Successfully added {sheet_name} sheet from main_output.xlsx")
#                 else:
#                     logger.warning(f"Sheet '{sheet_name}' not found in main_output.xlsx. Skipping...")
#                     st.warning(f"Sheet '{sheet_name}' not found in main_output.xlsx. Skipping...")
#         except Exception as e:
#             logger.error(f"Error processing main_output.xlsx: {str(e)}")
#             st.warning(f"Error processing main_output.xlsx: {str(e)}. Skipping remaining sheets...")
#     else:
#         logger.warning(f"{output_files[1]} not found. Skipping...")
#         st.warning(f"{output_files[1]} not found. Skipping...")

#     # Add watsonx_processed_results_<timestamp>.xlsx data
#     watsonx_files = [f for f in os.listdir('.') if re.match(r'watsonx_processed_results_\d{8}_\d{6}\.xlsx', f)]
#     if watsonx_files:
#         latest_watsonx_file = max(watsonx_files, key=lambda x: os.path.getmtime(x))  # Get the latest file
#         try:
#             watsonx_wb = openpyxl.load_workbook(latest_watsonx_file)
#             for sheet_name in watsonx_wb.sheetnames:
#                 if sheet_name == "Modified Data":  # Target the specific sheet from your screenshot
#                     ws = watsonx_wb[sheet_name]
#                     new_ws = workbook.add_worksheet(sheet_name)
#                     for row in ws.rows:
#                         for cell in row:
#                             new_ws.write(cell.row - 1, cell.column - 1, cell.value)
#                     logger.info(f"Successfully added {sheet_name} sheet from {latest_watsonx_file}")
#                 else:
#                     logger.warning(f"Skipping sheet '{sheet_name}' from {latest_watsonx_file} as it is not 'Modified Data'")
#             os.remove(latest_watsonx_file)  # Optional: Remove the processed file to avoid clutter
#         except Exception as e:
#             logger.error(f"Error processing {latest_watsonx_file}: {str(e)}")
#             st.warning(f"Error processing {latest_watsonx_file}: {str(e)}. Skipping...")
#     else:
#         logger.warning("No watsonx_processed_results_*.xlsx files found. Skipping...")
#         st.warning("No watsonx_processed_results_*.xlsx files found. Skipping...")

#     # Save and close the Excel file
#     try:
#         output.save()
#         logger.info("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#         st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
        
#         # Verify file integrity and offer download
#         with open("final_output.xlsx", "rb") as f:
#             file_content = f.read()
#         st.download_button(
#             label="Download Final Excel",
#             data=file_content,
#             file_name="final_output.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )
#     except Exception as e:
#         logger.error(f"Error saving or reading final_output.xlsx for download: {str(e)}")
#         st.error("Failed to generate a valid Excel file. Please check the logs.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs (e.g., ncr_output.xlsx, main_output.xlsx, watsonx_processed_results_<timestamp>.xlsx), then click 'Combine All Outputs'.")



# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Define formats at module level
# title_format = None
# header_format = None
# subheader_format = None
# cell_format = None
# site_format = None

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your applications
# apps = {
#     "NCR App": "ncr.py",
#     "Report App": "appp.py",
#     "Project Main": "project_main.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])

# # Function to generate NCR Excel sheet
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None, worksheet=None):
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     if workbook is None:
#         output = io.BytesIO()
#         writer = pd.ExcelWriter(output, engine='xlsxwriter')
#         local_workbook = writer.book
#         local_worksheet = local_workbook.add_worksheet('NCR Report')
#     else:
#         local_workbook = workbook
#         local_worksheet = worksheet or local_workbook.add_worksheet('NCR Report')

#     if title_format is None:
#         title_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'fg_color': 'yellow', 'border': 1, 'font_size': 12})
#     if header_format is None:
#         header_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
#     if subheader_format is None:
#         subheader_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#     if cell_format is None:
#         cell_format = local_workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
#     if site_format is None:
#         site_format = local_workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1})
    
#     local_worksheet.set_column('A:A', 20)
#     local_worksheet.set_column('B:H', 12)
    
#     resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#     open_data = combined_result.get("NCR open beyond 21 days", {})
#     if not isinstance(resolved_data, dict) or "error" in resolved_data:
#         resolved_data = {"Sites": {}}
#     if not isinstance(open_data, dict) or "error" in open_data:
#         open_data = {"Sites": {}}
        
#     resolved_sites = resolved_data.get("Sites", {})
#     open_sites = open_data.get("Sites", {})
    
#     standard_sites = [
#         "Veridia-Club", "Veridia- Tower 01", "Veridia- Tower 02", "Veridia- Tower 03",
#         "Veridia- Tower 04", "Veridia- Tower 05", "Veridia- Tower 06", "Veridia- Tower 07",
#         "Veridia-Commercial", "External Development"
#     ]
    
#     def normalize_site_name(site):
#         if site in ["Veridia-Club", "Veridia-Commercial"]:
#             return site
#         match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#         if match:
#             num = match.group(1).zfill(2)
#             return f"Veridia- Tower {num}"
#         return site

#     site_mapping = {k: normalize_site_name(k) for k in (resolved_sites.keys() | open_sites.keys())}
#     sorted_sites = sorted(standard_sites)
    
#     local_worksheet.merge_range('A1:H1', report_title, title_format)
#     row = 1
#     local_worksheet.write(row, 0, 'Site', header_format)
#     local_worksheet.merge_range(row, 1, row, 3, 'NCR resolved beyond 21 days', header_format)
#     local_worksheet.merge_range(row, 4, row, 6, 'NCR open beyond 21 days', header_format)
#     local_worksheet.write(row, 7, 'Total', header_format)
    
#     row = 2
#     categories = ['Civil Finishing', 'MEP', 'Structure']
#     local_worksheet.write(row, 0, '', header_format)
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+1, cat, subheader_format)
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+4, cat, subheader_format)
#     local_worksheet.write(row, 7, '', header_format)
    
#     category_map = {'Civil Finishing': 'FW', 'MEP': 'MEP', 'Structure': 'SW'}
#     row = 3
#     site_totals = {}
    
#     for site in sorted_sites:
#         local_worksheet.write(row, 0, site, site_format)
#         original_resolved_key = next((k for k, v in site_mapping.items() if v == site), None)
#         original_open_key = next((k for k, v in site_mapping.items() if v == site), None)
#         site_total = 0
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_resolved_key and original_resolved_key in resolved_sites:
#                 value = resolved_sites[original_resolved_key].get(json_cat, 0)
#             local_worksheet.write(row, i+1, value, cell_format)
#             site_total += value
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_open_key and original_open_key in open_sites:
#                 value = open_sites[original_open_key].get(json_cat, 0)
#             local_worksheet.write(row, i+4, value, cell_format)
#             site_total += value
#         local_worksheet.write(row, 7, site_total, cell_format)
#         site_totals[site] = site_total
#         row += 1
    
#     if workbook is None:
#         writer.close()  # Use close() instead of save()
#         output.seek(0)
#         return output
#     return None

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#         "updated_progress_data.xlsx",
#     ]
    
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Process ncr_output.xlsx
#     if os.path.exists(output_files[0]):
#         try:
#             ncr_wb = openpyxl.load_workbook(output_files[0])
#             ncr_ws = ncr_wb['NCR Report']
#             data = [list(row) for row in ncr_ws.iter_rows(min_row=4, values_only=True)]
#             combined_result = {"NCR resolved beyond 21 days": {"Sites": {}}, "NCR open beyond 21 days": {"Sites": {}}}
#             sites = [row[0] for row in data if row[0]]
#             for i, site in enumerate(sites):
#                 resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][1:4]) if v is not None and v != 0}
#                 open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][4:7]) if v is not None and v != 0}
#                 if resolved_data:
#                     combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#                 if open_data:
#                     combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
#             generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
#             logger.info("Successfully added NCR Report sheet to final_output.xlsx")
#         except Exception as e:
#             logger.error(f"Error processing {output_files[0]}: {str(e)}")
#             st.warning(f"Error processing {output_files[0]}: {str(e)}. Skipping NCR Report...")
#     else:
#         logger.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")
#         st.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")

#     # Process main_output.xlsx
#     if os.path.exists(output_files[1]):
#         try:
#             main_wb = openpyxl.load_workbook(output_files[1])
#             st.write(f"Main sheets: {main_wb.sheetnames}")  # Debug: Check available sheets
#             for sheet_name in ['Raw Data']:  # Adjust sheet name based on debug output
#                 if sheet_name in main_wb.sheetnames:
#                     ws = main_wb[sheet_name]
#                     new_ws = workbook.add_worksheet(sheet_name)
#                     for row in ws.rows:
#                         for cell in row:
#                             new_ws.write(cell.row - 1, cell.column - 1, cell.value)
#                     logger.info(f"Successfully added {sheet_name} sheet from {output_files[1]}")
#                 else:
#                     logger.warning(f"Sheet '{sheet_name}' not found in {output_files[1]}. Skipping...")
#                     st.warning(f"Sheet '{sheet_name}' not found in {output_files[1]}. Skipping...")
#         except Exception as e:
#             logger.error(f"Error processing {output_files[1]}: {str(e)}")
#             st.warning(f"Error processing {output_files[1]}: {str(e)}. Skipping remaining sheets...")
#     else:
#         logger.warning(f"{output_files[1]} not found. Skipping...")
#         st.warning(f"{output_files[1]} not found. Skipping...")

#     # Process updated_progress_data.xlsx
#     if os.path.exists(output_files[2]):
#         try:
#             progress_wb = openpyxl.load_workbook(output_files[2])
#             for sheet_name in progress_wb.sheetnames:
#                 if sheet_name == "Modified Data":
#                     ws = progress_wb[sheet_name]
#                     new_ws = workbook.add_worksheet(sheet_name)
#                     for row in ws.rows:
#                         for cell in row:
#                             new_ws.write(cell.row - 1, cell.column - 1, cell.value)
#                     logger.info(f"Successfully added {sheet_name} sheet from {output_files[2]}")
#                 else:
#                     logger.warning(f"Skipping sheet '{sheet_name}' from {output_files[2]} as it is not 'Modified Data'")
#         except Exception as e:
#             logger.error(f"Error processing {output_files[2]}: {str(e)}")
#             st.warning(f"Error processing {output_files[2]}: {str(e)}. Skipping...")
#     else:
#         logger.warning(f"{output_files[2]} not found. Please run Project Main first.")
#         st.warning(f"{output_files[2]} not found. Please run Project Main first.")

#     # Save and close the Excel file
#     try:
#         output.close()  # Changed from output.save() to output.close()
#         if len(workbook.sheetnames) > 0:
#             logger.info("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#             st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#             with open("final_output.xlsx", "rb") as f:
#                 file_content = f.read()
#             st.download_button(
#                 label="Download Final Excel",
#                 data=file_content,
#                 file_name="final_output.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )
#         else:
#             logger.warning("No sheets added to final_output.xlsx. Removing empty file.")
#             st.warning("No sheets added to final_output.xlsx. Please ensure input files are generated.")
#             if os.path.exists("final_output.xlsx"):
#                 try:
#                     os.remove("final_output.xlsx")
#                 except PermissionError as e:
#                     logger.warning(f"Could not remove final_output.xlsx due to: {str(e)}. Please close the file if open.")
#                     st.warning(f"Could not remove final_output.xlsx. Please close the file if open and try again.")
#     except Exception as e:
#         logger.error(f"Error saving or reading final_output.xlsx for download: {str(e)}")
#         st.error(f"Failed to generate a valid Excel file. Please check the logs: {str(e)}")
#         if os.path.exists("final_output.xlsx"):
#             try:
#                 os.remove("final_output.xlsx")
#             except PermissionError as e:
#                 logger.warning(f"Could not remove final_output.xlsx due to: {str(e)}. Please close the file if open.")
#                 st.warning(f"Could not remove final_output.xlsx. Please close the file if open and try again.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs (e.g., ncr_output.xlsx, main_output.xlsx, updated_progress_data.xlsx), then click 'Combine All Outputs'.")

# import streamlit as st
# import subprocess
# import pandas as pd
# import os
# import xlsxwriter
# import openpyxl
# import io
# import re
# import logging
# from datetime import datetime

# # Configure logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Define formats at module level
# title_format = None
# header_format = None
# subheader_format = None
# cell_format = None
# site_format = None

# # Apply custom CSS for a polished look and centered buttons
# st.markdown(
#     """
#     <style>
#     .main {
#         background-color: #ffffff;
#         padding: 20px;
#         border-radius: 10px;
#         box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
#         max-width: 800px;
#         margin: 0 auto;
#     }
#     .stApp > header {
#         background-color: #2c3e50;
#         color: white;
#     }
#     .css-1d391kg { /* Target the title */
#         color: #ffffff;
#         text-align: center;
#         font-size: 2.5em;
#         font-weight: bold;
#         padding: 10px;
#     }
#     .st-bv { /* Target subheaders */
#         color: #34495e;
#         text-align: center;
#         font-size: 1.5em;
#         margin-top: 20px;
#     }
#     .stButton>button {
#         display: block;
#         margin: 0 auto 15px auto;
#         width: 250px;
#         background-color: #3498db;
#         color: white;
#         border: none;
#         padding: 12px 24px;
#         border-radius: 8px;
#         font-size: 1.1em;
#         cursor: pointer;
#         transition: background-color 0.3s, transform 0.2s;
#         box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
#     }
#     .stButton>button:hover {
#         background-color: #2980b9;
#         transform: translateY(-2px);
#     }
#     .stSuccess, .stWarning, .stError {
#         text-align: center;
#         padding: 10px;
#         border-radius: 5px;
#         margin: 10px 0;
#     }
#     .stText {
#         text-align: center;
#         color: #7f8c8d;
#         font-size: 1em;
#         margin-top: 20px;
#     }
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# st.title("Wave Infra Application Launcher")

# # List of your applications
# apps = {
#     "NCR App": "ncr.py",
#     "Report App": "test.py",
#     "Project Main": "project_main.py",
# }

# # Display buttons to launch each app
# st.subheader("Launch Applications")
# for app_name, app_file in apps.items():
#     if st.button(f"Run {app_name}"):
#         st.write(f"Launching {app_name}...")
#         subprocess.Popen(["streamlit", "run", app_file])

# # Function to generate NCR Excel sheet
# def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None, worksheet=None):
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     if workbook is None:
#         output = io.BytesIO()
#         writer = pd.ExcelWriter(output, engine='xlsxwriter')
#         local_workbook = writer.book
#         local_worksheet = local_workbook.add_worksheet('NCR Report')
#     else:
#         local_workbook = workbook
#         local_worksheet = worksheet or local_workbook.add_worksheet('NCR Report')

#     if title_format is None:
#         title_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'fg_color': 'yellow', 'border': 1, 'font_size': 12})
#     if header_format is None:
#         header_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
#     if subheader_format is None:
#         subheader_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
#     if cell_format is None:
#         cell_format = local_workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
#     if site_format is None:
#         site_format = local_workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1})
    
#     local_worksheet.set_column('A:A', 20)
#     local_worksheet.set_column('B:H', 12)
    
#     resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
#     open_data = combined_result.get("NCR open beyond 21 days", {})
#     if not isinstance(resolved_data, dict) or "error" in resolved_data:
#         resolved_data = {"Sites": {}}
#     if not isinstance(open_data, dict) or "error" in open_data:
#         open_data = {"Sites": {}}
        
#     resolved_sites = resolved_data.get("Sites", {})
#     open_sites = open_data.get("Sites", {})
    
#     standard_sites = [
#         "Veridia-Club", "Veridia- Tower 01", "Veridia- Tower 02", "Veridia- Tower 03",
#         "Veridia- Tower 04", "Veridia- Tower 05", "Veridia- Tower 06", "Veridia- Tower 07",
#         "Veridia-Commercial", "External Development"
#     ]
    
#     def normalize_site_name(site):
#         if site in ["Veridia-Club", "Veridia-Commercial"]:
#             return site
#         match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
#         if match:
#             num = match.group(1).zfill(2)
#             return f"Veridia- Tower {num}"
#         return site

#     site_mapping = {k: normalize_site_name(k) for k in (resolved_sites.keys() | open_sites.keys())}
#     sorted_sites = sorted(standard_sites)
    
#     local_worksheet.merge_range('A1:H1', report_title, title_format)
#     row = 1
#     local_worksheet.write(row, 0, 'Site', header_format)
#     local_worksheet.merge_range(row, 1, row, 3, 'NCR resolved beyond 21 days', header_format)
#     local_worksheet.merge_range(row, 4, row, 6, 'NCR open beyond 21 days', header_format)
#     local_worksheet.write(row, 7, 'Total', header_format)
    
#     row = 2
#     categories = ['Civil Finishing', 'MEP', 'Structure']
#     local_worksheet.write(row, 0, '', header_format)
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+1, cat, subheader_format)
#     for i, cat in enumerate(categories):
#         local_worksheet.write(row, i+4, cat, subheader_format)
#     local_worksheet.write(row, 7, '', header_format)
    
#     category_map = {'Civil Finishing': 'FW', 'MEP': 'MEP', 'Structure': 'SW'}
#     row = 3
#     site_totals = {}
    
#     for site in sorted_sites:
#         local_worksheet.write(row, 0, site, site_format)
#         original_resolved_key = next((k for k, v in site_mapping.items() if v == site), None)
#         original_open_key = next((k for k, v in site_mapping.items() if v == site), None)
#         site_total = 0
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_resolved_key and original_resolved_key in resolved_sites:
#                 value = resolved_sites[original_resolved_key].get(json_cat, 0)
#             local_worksheet.write(row, i+1, value, cell_format)
#             site_total += value
#         for i, (display_cat, json_cat) in enumerate(category_map.items()):
#             value = 0
#             if original_open_key and original_open_key in open_sites:
#                 value = open_sites[original_open_key].get(json_cat, 0)
#             local_worksheet.write(row, i+4, value, cell_format)
#             site_total += value
#         local_worksheet.write(row, 7, site_total, cell_format)
#         site_totals[site] = site_total
#         row += 1
    
#     if workbook is None:
#         writer.close()  # Use close() instead of save()
#         output.seek(0)
#         return output
#     return None

# # Function to combine Excel outputs with preserved structure
# def combine_excel_outputs():
#     global title_format, header_format, subheader_format, cell_format, site_format
    
#     output_files = [
#         "ncr_output.xlsx",
#         "main_output.xlsx",
#         "updated_progress_data.xlsx",
#     ]
    
#     output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
#     workbook = output.book

#     # Process ncr_output.xlsx
#     if os.path.exists(output_files[0]):
#         try:
#             ncr_wb = openpyxl.load_workbook(output_files[0])
#             ncr_ws = ncr_wb['NCR Report']
#             data = [list(row) for row in ncr_ws.iter_rows(min_row=4, values_only=True)]
#             combined_result = {"NCR resolved beyond 21 days": {"Sites": {}}, "NCR open beyond 21 days": {"Sites": {}}}
#             sites = [row[0] for row in data if row[0]]
#             for i, site in enumerate(sites):
#                 resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][1:4]) if v is not None and v != 0}
#                 open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][4:7]) if v is not None and v != 0}
#                 if resolved_data:
#                     combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
#                 if open_data:
#                     combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
#             generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
#             logger.info("Successfully added NCR Report sheet to final_output.xlsx")
#         except Exception as e:
#             logger.error(f"Error processing {output_files[0]}: {str(e)}")
#             st.warning(f"Error processing {output_files[0]}: {str(e)}. Skipping NCR Report...")
#     else:
#         logger.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")
#         st.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")

#     # Process main_output.xlsx
#     if os.path.exists(output_files[1]):
#         try:
#             main_wb = openpyxl.load_workbook(output_files[1])
#             st.write(f"Main sheets: {main_wb.sheetnames}")  # Debug: Check available sheets
#             for sheet_name in ['Raw Data']:  # Matches the updated sheet name from appp.py
#                 if sheet_name in main_wb.sheetnames:
#                     ws = main_wb[sheet_name]
#                     new_ws = workbook.add_worksheet(sheet_name)
#                     for row in ws.rows:
#                         for cell in row:
#                             new_ws.write(cell.row - 1, cell.column - 1, cell.value)
#                     logger.info(f"Successfully added {sheet_name} sheet from {output_files[1]}")
#                 else:
#                     logger.warning(f"Sheet '{sheet_name}' not found in {output_files[1]}. Skipping...")
#                     st.warning(f"Sheet '{sheet_name}' not found in {output_files[1]}. Skipping...")
#         except Exception as e:
#             logger.error(f"Error processing {output_files[1]}: {str(e)}")
#             st.warning(f"Error processing {output_files[1]}: {str(e)}. Skipping remaining sheets...")
#     else:
#         logger.warning(f"{output_files[1]} not found. Skipping...")
#         st.warning(f"{output_files[1]} not found. Skipping...")

#     # Process updated_progress_data.xlsx
#     if os.path.exists(output_files[2]):
#         try:
#             progress_wb = openpyxl.load_workbook(output_files[2])
#             for sheet_name in progress_wb.sheetnames:
#                 if sheet_name == "Modified Data":
#                     ws = progress_wb[sheet_name]
#                     new_ws = workbook.add_worksheet(sheet_name)
#                     for row in ws.rows:
#                         for cell in row:
#                             new_ws.write(cell.row - 1, cell.column - 1, cell.value)
#                     logger.info(f"Successfully added {sheet_name} sheet from {output_files[2]}")
#                 else:
#                     logger.warning(f"Skipping sheet '{sheet_name}' from {output_files[2]} as it is not 'Modified Data'")
#         except Exception as e:
#             logger.error(f"Error processing {output_files[2]}: {str(e)}")
#             st.warning(f"Error processing {output_files[2]}: {str(e)}. Skipping...")
#     else:
#         logger.warning(f"{output_files[2]} not found. Please run Project Main first.")
#         st.warning(f"{output_files[2]} not found. Please run Project Main first.")

#     # Save and close the Excel file
#     try:
#         output.close()  # Changed from output.save() to output.close()
#         if len(workbook.sheetnames) > 0:
#             logger.info("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#             st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
#             with open("final_output.xlsx", "rb") as f:
#                 file_content = f.read()
#             st.download_button(
#                 label="Download Final Excel",
#                 data=file_content,
#                 file_name="final_output.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )
#         else:
#             logger.warning("No sheets added to final_output.xlsx. Removing empty file.")
#             st.warning("No sheets added to final_output.xlsx. Please ensure input files are generated.")
#             if os.path.exists("final_output.xlsx"):
#                 try:
#                     os.remove("final_output.xlsx")
#                 except PermissionError as e:
#                     logger.warning(f"Could not remove final_output.xlsx due to: {str(e)}. Please close the file if open.")
#                     st.warning(f"Could not remove final_output.xlsx. Please close the file if open and try again.")
#     except Exception as e:
#         logger.error(f"Error saving or reading final_output.xlsx for download: {str(e)}")
#         st.error(f"Failed to generate a valid Excel file. Please check the logs: {str(e)}")
#         if os.path.exists("final_output.xlsx"):
#             try:
#                 os.remove("final_output.xlsx")
#             except PermissionError as e:
#                 logger.warning(f"Could not remove final_output.xlsx due to: {str(e)}. Please close the file if open.")
#                 st.warning(f"Could not remove final_output.xlsx. Please close the file if open and try again.")

# # Button to generate the final Excel
# st.subheader("Generate Final Excel")
# if st.button("Combine All Outputs"):
#     combine_excel_outputs()

# st.write("Note: Run each app first to generate their outputs (e.g., ncr_output.xlsx, main_output.xlsx, updated_progress_data.xlsx), then click 'Combine All Outputs'.")



import streamlit as st
import subprocess
import pandas as pd
import os
import xlsxwriter
import openpyxl
import io
import re
import logging
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Define formats at module level
title_format = None
header_format = None
subheader_format = None
cell_format = None
site_format = None

# Apply custom CSS for a polished look and centered buttons
st.markdown(
    """
    <style>
    .main {
        background-color: #ffffff;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
        max-width: 800px;
        margin: 0 auto;
    }
    .stApp > header {
        background-color: #2c3e50;
        color: white;
    }
    .css-1d391kg { /* Target the title */
        color: #ffffff;
        text-align: center;
        font-size: 2.5em;
        font-weight: bold;
        padding: 10px;
    }
    .st-bv { /* Target subheaders */
        color: #34495e;
        text-align: center;
        font-size: 1.5em;
        margin-top: 20px;
    }
    .stButton>button {
        display: block;
        margin: 0 auto 15px auto;
        width: 250px;
        background-color: #3498db;
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        font-size: 1.1em;
        cursor: pointer;
        transition: background-color 0.3s, transform 0.2s;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    }
    .stButton>button:hover {
        background-color: #2980b9;
        transform: translateY(-2px);
    }
    .stSuccess, .stWarning, .stError {
        text-align: center;
        padding: 10px;
        border-radius: 5px;
        margin: 10px 0;
    }
    .stText {
        text-align: center;
        color: #7f8c8d;
        font-size: 1em;
        margin-top: 20px;
    }
    </style>
    """,
    unsafe_allow_html=True
)

st.title("Wave Infra Application Launcher")

# List of your applications
apps = {
    "NCR App": "ncr.py",
    "Report App": "test.py",
    "Project Main": "structure_and_finishing_main.py",
}

# Display buttons to launch each app
st.subheader("Launch Applications")
for app_name, app_file in apps.items():
    if st.button(f"Run {app_name}"):
        st.write(f"Launching {app_name}...")
        subprocess.Popen(["streamlit", "run", app_file])

# Function to generate NCR Excel sheet
def generate_consolidated_ncr_excel(combined_result, report_title="NCR", workbook=None, worksheet=None):
    global title_format, header_format, subheader_format, cell_format, site_format
    
    if workbook is None:
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        local_workbook = writer.book
        local_worksheet = local_workbook.add_worksheet('NCR Report')
    else:
        local_workbook = workbook
        local_worksheet = worksheet or local_workbook.add_worksheet('NCR Report')

    if title_format is None:
        title_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'fg_color': 'yellow', 'border': 1, 'font_size': 12})
    if header_format is None:
        header_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
    if subheader_format is None:
        subheader_format = local_workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    if cell_format is None:
        cell_format = local_workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    if site_format is None:
        site_format = local_workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1})
    
    local_worksheet.set_column('A:A', 20)
    local_worksheet.set_column('B:H', 12)
    
    resolved_data = combined_result.get("NCR resolved beyond 21 days", {})
    open_data = combined_result.get("NCR open beyond 21 days", {})
    if not isinstance(resolved_data, dict) or "error" in resolved_data:
        resolved_data = {"Sites": {}}
    if not isinstance(open_data, dict) or "error" in open_data:
        open_data = {"Sites": {}}
        
    resolved_sites = resolved_data.get("Sites", {})
    open_sites = open_data.get("Sites", {})
    
    standard_sites = [
        "Veridia-Club", "Veridia- Tower 01", "Veridia- Tower 02", "Veridia- Tower 03",
        "Veridia- Tower 04", "Veridia- Tower 05", "Veridia- Tower 06", "Veridia- Tower 07",
        "Veridia-Commercial", "External Development"
    ]
    
    def normalize_site_name(site):
        if site in ["Veridia-Club", "Veridia-Commercial"]:
            return site
        match = re.search(r'(?:tower|t)[- ]?(\d+)', site, re.IGNORECASE)
        if match:
            num = match.group(1).zfill(2)
            return f"Veridia- Tower {num}"
        return site

    site_mapping = {k: normalize_site_name(k) for k in (resolved_sites.keys() | open_sites.keys())}
    sorted_sites = sorted(standard_sites)
    
    local_worksheet.merge_range('A1:H1', report_title, title_format)
    row = 1
    local_worksheet.write(row, 0, 'Site', header_format)
    local_worksheet.merge_range(row, 1, row, 3, 'NCR resolved beyond 21 days', header_format)
    local_worksheet.merge_range(row, 4, row, 6, 'NCR open beyond 21 days', header_format)
    local_worksheet.write(row, 7, 'Total', header_format)
    
    row = 2
    categories = ['Civil Finishing', 'MEP', 'Structure']
    local_worksheet.write(row, 0, '', header_format)
    for i, cat in enumerate(categories):
        local_worksheet.write(row, i+1, cat, subheader_format)
    for i, cat in enumerate(categories):
        local_worksheet.write(row, i+4, cat, subheader_format)
    local_worksheet.write(row, 7, '', header_format)
    
    category_map = {'Civil Finishing': 'FW', 'MEP': 'MEP', 'Structure': 'SW'}
    row = 3
    site_totals = {}
    
    for site in sorted_sites:
        local_worksheet.write(row, 0, site, site_format)
        original_resolved_key = next((k for k, v in site_mapping.items() if v == site), None)
        original_open_key = next((k for k, v in site_mapping.items() if v == site), None)
        site_total = 0
        for i, (display_cat, json_cat) in enumerate(category_map.items()):
            value = 0
            if original_resolved_key and original_resolved_key in resolved_sites:
                value = resolved_sites[original_resolved_key].get(json_cat, 0)
            local_worksheet.write(row, i+1, value, cell_format)
            site_total += value
        for i, (display_cat, json_cat) in enumerate(category_map.items()):
            value = 0
            if original_open_key and original_open_key in open_sites:
                value = open_sites[original_open_key].get(json_cat, 0)
            local_worksheet.write(row, i+4, value, cell_format)
            site_total += value
        local_worksheet.write(row, 7, site_total, cell_format)
        site_totals[site] = site_total
        row += 1
    
    if workbook is None:
        writer.close()  # Use close() instead of save()
        output.seek(0)
        return output
    return None

# Function to combine Excel outputs with preserved structure
def combine_excel_outputs():
    global title_format, header_format, subheader_format, cell_format, site_format
    
    output_files = [
        "ncr_output.xlsx",
        "main_output.xlsx",
        "updated_progress_data.xlsx",
    ]
    
    output = pd.ExcelWriter("final_output.xlsx", engine='xlsxwriter')
    workbook = output.book

    # Process ncr_output.xlsx
    if os.path.exists(output_files[0]):
        try:
            ncr_wb = openpyxl.load_workbook(output_files[0])
            ncr_ws = ncr_wb['NCR Report']
            data = [list(row) for row in ncr_ws.iter_rows(min_row=4, values_only=True)]
            combined_result = {"NCR resolved beyond 21 days": {"Sites": {}}, "NCR open beyond 21 days": {"Sites": {}}}
            sites = [row[0] for row in data if row[0]]
            for i, site in enumerate(sites):
                resolved_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][1:4]) if v is not None and v != 0}
                open_data = {k: v for k, v in zip(['FW', 'MEP', 'SW'], data[i][4:7]) if v is not None and v != 0}
                if resolved_data:
                    combined_result["NCR resolved beyond 21 days"]["Sites"][site] = resolved_data
                if open_data:
                    combined_result["NCR open beyond 21 days"]["Sites"][site] = open_data
            generate_consolidated_ncr_excel(combined_result, report_title="NCR: April", workbook=workbook)
            logger.info("Successfully added NCR Report sheet to final_output.xlsx")
        except Exception as e:
            logger.error(f"Error processing {output_files[0]}: {str(e)}")
            st.warning(f"Error processing {output_files[0]}: {str(e)}. Skipping NCR Report...")
    else:
        logger.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")
        st.warning(f"{output_files[0]} not found. Skipping NCR Report generation...")

    # Process main_output.xlsx
    if os.path.exists(output_files[1]):
        try:
            main_wb = openpyxl.load_workbook(output_files[1])
            st.write(f"Main sheets: {main_wb.sheetnames}")  # Debug: Check available sheets
            for sheet_name in ['Activity Counts']:
                if sheet_name in main_wb.sheetnames:
                    # Read the DataFrame from the sheet to preserve pivot table structure
                    df = pd.read_excel(output_files[1], sheet_name=sheet_name, index_col=0)
                    # Write the DataFrame to the new workbook
                    df.to_excel(output, sheet_name=sheet_name, index=True)
                    logger.info(f"Successfully added {sheet_name} sheet from {output_files[1]}")
                else:
                    logger.warning(f"Sheet '{sheet_name}' not found in {output_files[1]}. Skipping...")
                    st.warning(f"Sheet '{sheet_name}' not found in {output_files[1]}. Skipping...")
        except Exception as e:
            logger.error(f"Error processing {output_files[1]}: {str(e)}")
            st.warning(f"Error processing {output_files[1]}: {str(e)}. Skipping remaining sheets...")
    else:
        logger.warning(f"{output_files[1]} not found. Skipping...")
        st.warning(f"{output_files[1]} not found. Skipping...")

    # Process updated_progress_data.xlsx
    if os.path.exists(output_files[2]):
        try:
            progress_wb = openpyxl.load_workbook(output_files[2])
            for sheet_name in progress_wb.sheetnames:
                if sheet_name == "Modified Data":
                    ws = progress_wb[sheet_name]
                    new_ws = workbook.add_worksheet(sheet_name)
                    for row in ws.rows:
                        for cell in row:
                            new_ws.write(cell.row - 1, cell.column - 1, cell.value)
                    logger.info(f"Successfully added {sheet_name} sheet from {output_files[2]}")
                else:
                    logger.warning(f"Skipping sheet '{sheet_name}' from {output_files[2]} as it is not 'Modified Data'")
        except Exception as e:
            logger.error(f"Error processing {output_files[2]}: {str(e)}")
            st.warning(f"Error processing {output_files[2]}: {str(e)}. Skipping...")
    else:
        logger.warning(f"{output_files[2]} not found. Please run Project Main first.")
        st.warning(f"{output_files[2]} not found. Please run Project Main first.")

    # Save and close the Excel file
    try:
        output.close()  # Changed from output.save() to output.close()
        if len(workbook.sheetnames) > 0:
            logger.info("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
            st.success("Final Excel file generated as 'final_output.xlsx' with multiple sheets!")
            with open("final_output.xlsx", "rb") as f:
                file_content = f.read()
            st.download_button(
                label="Download Final Excel",
                data=file_content,
                file_name="final_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            logger.warning("No sheets added to final_output.xlsx. Removing empty file.")
            st.warning("No sheets added to final_output.xlsx. Please ensure input files are generated.")
            if os.path.exists("final_output.xlsx"):
                try:
                    os.remove("final_output.xlsx")
                except PermissionError as e:
                    logger.warning(f"Could not remove final_output.xlsx due to: {str(e)}. Please close the file if open.")
                    st.warning(f"Could not remove final_output.xlsx. Please close the file if open and try again.")
    except Exception as e:
        logger.error(f"Error saving or reading final_output.xlsx for download: {str(e)}")
        st.error(f"Failed to generate a valid Excel file. Please check the logs: {str(e)}")
        if os.path.exists("final_output.xlsx"):
            try:
                os.remove("final_output.xlsx")
            except PermissionError as e:
                logger.warning(f"Could not remove final_output.xlsx due to: {str(e)}. Please close the file if open.")
                st.warning(f"Could not remove final_output.xlsx. Please close the file if open and try again.")

# Button to generate the final Excel
st.subheader("Generate Final Excel")
if st.button("Combine All Outputs"):
    combine_excel_outputs()

st.write("Note: Run each app first to generate their outputs (e.g., ncr_output.xlsx, main_output.xlsx, updated_progress_data.xlsx), then click 'Combine All Outputs'.")