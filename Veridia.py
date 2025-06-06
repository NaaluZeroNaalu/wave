import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string



tower2 = []
tower3 = []
tower4 = []
tower5 = []
tower6 = []
tower7 = []


WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [[{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()




def Tower2(sheet):
    st.write("Analyzing Veridia Tower 2")
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    tower2.append(1)
                else:
                    tower2.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")
                
    # for row in rows:

    #     for col in cols:
    #         col_idx = column_index_from_string(col)
    #         cell = sheet.cell(row=row, column=col_idx)
    #         value = cell.value
    #         color = get_fill_color(cell)
    #         # st.write(f"{col}{row}")
    #         # st.write(value)
    #         # st.write(color)
    #         if color == "Theme (9)" or color == "Theme (7)" or color == "Theme (5)" or color == "FF92D050" or color == "FFFFFF00":
    #             tower2.append(1)
    #         else:
    #             tower2.append(0)

def Tower3(sheet):
    st.write("Analyzing Veridia Tower 3")
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH']

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    tower3.append(1)
                else:
                    tower3.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")
                
            
    # for row in rows:

    #     for col in cols:
    #         col_idx = column_index_from_string(col)
    #         cell = sheet.cell(row=row, column=col_idx)
    #         value = cell.value
    #         color = get_fill_color(cell)
    #         # st.write(f"{col}{row}")
    #         # st.write(value)
    #         # st.write(color)
    #         if color == "Theme (9)" or color == "Theme (7)" or color == "Theme (5)" or color == "FF92D050" or color == "FFFFFF00":
    #             tower3.append(1)
    #         else:
    #             tower3.append(0)

def Tower4(sheet):
    st.write("Analyzing Veridia Tower 4")
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX', 'AZ', 'BB', 'BD', 'BF', 'BH', 'BJ', 'BL', 'BN', 'BP']
    
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    tower4.append(1)
                else:
                    tower4.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")
                

def Tower5(sheet):
    st.write("Analyzing Veridia Tower 5")
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['DC', 'DE', 'DG', 'DI', 'DK', 'DM', 'DO', 'DQ', 'DS', 'DU', 'DW', 'DY', 'EA', 'EC']

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    tower5.append(1)
                else:
                    tower5.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")
                

def Tower6(sheet):
    st.write("Analyzing Veridia Tower 6")
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['FK', 'FM', 'FO', 'FQ', 'FS', 'FU', 'FW', 'FY',  'GA', 'GC', 'GE', 'GG', 'GI', "GK"]

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    tower6.append(1)
                else:
                    tower6.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")
                

def Tower7(sheet):
    st.write("Analyzing Veridia Tower 7")

    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['EG', 'EI', 'EK', 'EM', 'EO', 'EQ', 'ES', 'EU', 'EW', 'EY', 'FA', 'FC', 'FE', 'FG']
    
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]  # Access cell using A1 notation
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb  # Get RGB color code
                # st.write(f"Cell {col}{row} color: {color}, value: {cell.value}")
                if color == "FF92D050":
                    tower7.append(1)
                else:
                    tower7.append(0)
            else:
                st.write(f"Cell {col}{row} has no solid fill color, value: {cell.value}")
                

def ProcessVeridia(exceldatas):
    wb = load_workbook(exceldatas)
    sheet_names = wb.sheetnames
    sheet_name = "Revised baseline with 60d NGT"

    sheet = wb[sheet_name]

    Tower2(sheet)
    st.divider()
    Tower3(sheet)
    st.divider()
    Tower4(sheet)
    st.divider()
    Tower5(sheet)
    st.divider()
    # st.write(test)
    Tower6(sheet)
    st.divider()
    Tower7(sheet)
    # st.write(tower2.count(1))
    # st.write(tower3.count(1))
    # st.write(tower4.count(1))
    # st.write(tower5.count(1))
    # st.write(tower6.count(1))
    # st.write(tower7.count(1))
    data = {
    "Project Name":["VERIDIA", "VERIDIA", "VERIDIA","VERIDIA","VERIDIA","VERIDIA"],
    "Tower": ["TOWER 2", "TOWER 3", "TOWER 4", "TOWER 5", "TOWER 6", "TOWER 7"],
    "Green (1)": [tower2.count(1), tower3.count(1), tower4.count(1), tower5.count(1), tower6.count(1), tower7.count(1)],
    "Non-Green (0)": [tower2.count(0), tower3.count(0), tower4.count(0), tower5.count(0), tower6.count(0), tower7.count(0)],
    "Finishing":[st.session_state.tower2_finishing,st.session_state.tower3_finishing,st.session_state.tower4_finishing,st.session_state.tower5_finishing,st.session_state.tower6_finishing,st.session_state.tower7_finishing]
}
    # st.table(data)
    df = pd.DataFrame(data)

    Ai_answer = generatePrompt(df)
    st.write(Ai_answer)
    st.write(f"Tower4 Finishing{st.session_state.tower4_finishing}")
    json_data = json.loads(Ai_answer)

    st.write(json_data)
    return json_data
