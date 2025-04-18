import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
import io

st.title("Excel File Reader with Month and Year Filter")


WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"


if 'processed_df' not in st.session_state:
    st.session_state.processed_df = None
if 'total_count_df' not in st.session_state:
    st.session_state.total_count_df = None
if 'selected_file_name' not in st.session_state:
    st.session_state.selected_file_name = None
if 'sheduledf' not in st.session_state:
    st.session_state.sheduledf = None
if 'shedule' not in st.session_state:
    st.session_state.shedule = None

def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY  # Define API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    if response.status_code != 200:
        st.error(f"Failed to get access token: {response.text}")
        return None
    return response.json()['access_token']

# Generate prompt for Watson API
def generatePrompt(json_datas):
    body = {
        "input": f"""
        Parse this JSON data, where each entry represents a task with 'Activity Name' and 'Finish_Month_Name' fields:
        {json_datas}
Count the number of occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)} for the selected year. 
Return a JSON object where:
- Keys are unique 'Activity Name' values.
- Values are dictionaries with keys as 'Finish_Month_Name' from {', '.join(selected_months)} and values as the count of occurrences.
- Include all months from {', '.join(selected_months)} for every 'Activity Name', setting the count to 0 if there are no occurrences in that month.
Example output:
{{
    "Install Windows": {{"Mar": 2, "Apr": 1, "May": 0}},
    "Paint Walls": {{"Mar": 1, "Apr": 3, "May": 0}}
}}
Return only the JSON object, no code, no explanation, just the formatted JSON.
        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,  # Define MODEL_ID
        "project_id": PROJECT_ID  # Define PROJECT_ID
    }
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    if not headers["Authorization"]:
        return "Error: No valid access token."
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)  # Define WATSONX_API_URL
    if response.status_code != 200:
        st.error(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    return response.json()['results'][0]['generated_text'].strip()

# Function to create chunks and process data
def createChunk(result_json, chunk_size=2000):
    num_rows = len(result_json)
    num_chunks = (num_rows + chunk_size - 1) // chunk_size  
    all_chunks = {} 
    temp = []
    st.write(f"Filtered rows in JSON format (split into {chunk_size}-row chunks):")
    for i in range(num_chunks):
        start_idx = i * chunk_size
        end_idx = min((i + 1) * chunk_size, num_rows)
        chunk_data = result_json[start_idx:end_idx]
        processed_data = json.loads(generatePrompt(chunk_data)) 
        temp.append(processed_data)  
        all_chunks[f"chunk_{i + 1}"] = chunk_data
    
    table_data = []
    all_months = set()
    for chunk in temp:
        for activity, months in chunk.items():
            all_months.update(months.keys())
    all_months = sorted(all_months, key=lambda x: pd.to_datetime(x, format='%b').month)
    
    for chunk in temp:
        for activity, months in chunk.items():
            row = {'Activity Name': activity}
            total_count = 0
            for month in all_months:
                count = months.get(month, 0)
                row[month] = count
                total_count += count
            row['Total Count'] = total_count
            table_data.append(row)
    
    df = pd.DataFrame(table_data)
    st.write("Activity Counts Table by Month")
    st.dataframe(df)

# Function to process Excel file
def process_file(file_stream):
    workbook = openpyxl.load_workbook(file_stream)
    activity_col_idx = 5  
    if "TOWER 4 FINISHING." in workbook.sheetnames:
        sheet_name = "TOWER 4 FINISHING."
        sheet = workbook[sheet_name]
        non_bold_rows = [
            row_idx for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16)
            if row[activity_col_idx].font and not row[activity_col_idx].font.b
        ]
        df = pd.read_excel(file_stream, sheet_name=sheet_name, skiprows=15)
        df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
                      'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                      'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
    
    elif "M7 T5" in workbook.sheetnames:
        sheet_name = "M7 T5"
        sheet = workbook[sheet_name]
        non_bold_rows = [
            row_idx for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16)
            if row[activity_col_idx].font and not row[activity_col_idx].font.b  # Fixed index
        ]
        df = pd.read_excel(file_stream, sheet_name=sheet_name)
        expected_columns = ['Module', 'Floor', 'Flat', 'Domain', 'Monthly Look', 'Activity ID', 'Activity Name', 
                           'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                           'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
        actual_cols = len(df.columns)
        if actual_cols >= len(expected_columns):
            df.columns = expected_columns + [f'Extra_{i}' for i in range(actual_cols - len(expected_columns))]
        else:
            df.columns = expected_columns[:actual_cols]
    else:
        return None

    required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
    df = df[[col for col in required_columns if col in df.columns]]
    df.index = df.index + 16
    df = df.loc[df.index.isin(non_bold_rows)]
    
    if 'Start' in df.columns:
        df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
    if 'Finish' in df.columns:
        df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)
        df['Finish_Year'] = df['Finish'].dt.year
        df['Finish_Month'] = df['Finish'].dt.month
        df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')
    
    return df

# Function to convert DataFrame to Excel
def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, sheet_name='Activity Counts')
    return output.getvalue()

# Function to get COS files
def get_cos_files():
    try:
        response = st.session_state.cos_client.list_objects_v2(Bucket="projectreport")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            st.warning("No .xlsx files found in the bucket 'projectreport'. Please ensure Excel files are uploaded.")
        return files
    except Exception as e:
        st.error(f"Error fetching COS files: {e}")
        return []

# Streamlit App
st.title("Excel File Activity Processor")

# Get files from COS
files = get_cos_files()

# Sidebar file selector
if files:
    selected_file_name = st.sidebar.selectbox("Select a file to process", files)
    
    # Process file only if the selected file has changed
    if selected_file_name and st.session_state.selected_file_name != selected_file_name:
        try:
            response = st.session_state.cos_client.get_object(Bucket="projectreport", Key=selected_file_name)
            st.session_state.processed_df = process_file(io.BytesIO(response['Body'].read()))
            st.session_state.selected_file_name = selected_file_name
            st.success(f"Processed file: {selected_file_name}")
        except Exception as e:
            st.error(f"Error processing file {selected_file_name}: {e}")
            st.session_state.processed_df = None

# Data display and filtering
if st.session_state.get("processed_df") is not None:
    df = st.session_state.processed_df

    available_years = sorted(df['Finish_Year'].dropna().unique())
    available_months = sorted(df['Finish_Month_Name'].dropna().unique())

    selected_year = st.sidebar.selectbox('Select Year', available_years, index=len(available_years)-1)
    selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

    # Apply filtering on cached DataFrame
    filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]

    st.write(f"Filtered rows for {', '.join(selected_months)} {selected_year}")
    st.write(filtered_df)
    st.write(f"Number of rows: {len(filtered_df)}")

    result = filtered_df[['Activity Name', 'Finish_Month_Name']]
    result_json = result.to_dict(orient='records')
    st.write(pd.DataFrame(result_json))

    mar_count = len(result[(result['Activity Name'] == 'EL-Third Fix ') & (result['Finish_Month_Name'] == 'Mar')])
    st.write(f"Number of 'EL-Third Fix' in 'Mar': {mar_count}")

    month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    activity_month_counts = pd.pivot_table(
        filtered_df,
        values='Activity ID',
        index='Activity Name',
        columns='Finish_Month_Name',
        aggfunc='count',
        fill_value=0
    )
    existing_months = [m for m in month_order if m in activity_month_counts.columns]
    activity_month_counts = activity_month_counts.reindex(columns=existing_months)
    activity_month_counts['Total Count'] = activity_month_counts.sum(axis=1)

    st.session_state.sheduledf = activity_month_counts
    st.session_state.shedule = to_excel(activity_month_counts)

    if st.button('Count The activity'):
        createChunk(result_json)
else:
    st.warning("No processed data available. Please select and process a file.")