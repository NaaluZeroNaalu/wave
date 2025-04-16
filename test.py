# import streamlit as st
# import pandas as pd
# import requests
# import json
# import openpyxl
# import time
# import math
# from io import BytesIO

# st.title("Excel File Reader with Month and Year Filter")


# WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
# MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
# PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
# API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"


# if 'processed_df' not in st.session_state:
#     st.session_state.processed_df = None

# if 'total_count_df' not in st.session_state:
#     st.session_state.total_count_df = None

# # Function to get access token
# def GetAccesstoken():
#     auth_url = "https://iam.cloud.ibm.com/identity/token"
    
#     headers = {
#         "Content-Type": "application/x-www-form-urlencoded",
#         "Accept": "application/json"
#     }
    
#     data = {
#         "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
#         "apikey": API_KEY
#     }
#     response = requests.post(auth_url, headers=headers, data=data)
    
#     if response.status_code != 200:
#         st.write(f"Failed to get access token: {response.text}")
#         return None
#     else:
#         token_info = response.json()
#         return token_info['access_token']

# # Generate prompt for Watson API
# def generatePrompt(json_datas):
#     body = {
#         "input": f"""
#         Parse this JSON data, where each entry represents a task with 'Activity Name' and 'Finish_Month_Name' fields:
#         {json_datas}
# Count the number of occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)} for the selected year. 
# Return a JSON object where:
# - Keys are unique 'Activity Name' values.
# - Values are dictionaries with keys as 'Finish_Month_Name' from {', '.join(selected_months)} and values as the count of occurrences.
# - Include all months from {', '.join(selected_months)} for every 'Activity Name', setting the count to 0 if there are no occurrences in that month.
# Example output:
# {{
#     "Install Windows": {{"Mar": 2, "Apr": 1, "May": 0}},
#     "Paint Walls": {{"Mar": 1, "Apr": 3, "May": 0}}
# }}
# Return only the JSON object, no code, no explanation, just the formatted JSON.

#         """, 
#         "parameters": {
#             "decoding_method": "greedy",
#             "max_new_tokens": 8100,
#             "min_new_tokens": 0,
#             "stop_sequences": [";"],
#             "repetition_penalty": 1.05,
#             "temperature": 0.5
#         },
#         "model_id": MODEL_ID,
#         "project_id": PROJECT_ID
#     }
    
#     headers = {
#         "Accept": "application/json",
#         "Content-Type": "application/json",
#         "Authorization": f"Bearer {GetAccesstoken()}"
#     }
    
#     if not headers["Authorization"]:
#         return "Error: No valid access token."
    
#     response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
#     if response.status_code != 200:
#         st.write(f"Failed to generate prompt: {response.text}")
#         return "Error generating prompt"
#     # st.write(json_datas)
#     return response.json()['results'][0]['generated_text'].strip()

# # Function to create chunks and process data
# def createChunk(result_json, chunk_size=2000):
#     num_rows = len(result_json)
#     num_chunks = (num_rows + chunk_size - 1) // chunk_size  
#     all_chunks = {} 
#     temp = []

#     st.write(f"Filtered rows in JSON format (split into {chunk_size}-row chunks):")
 
#     for i in range(num_chunks):
#         start_idx = i * chunk_size
#         end_idx = min((i + 1) * chunk_size, num_rows)
#         chunk_data = result_json[start_idx:end_idx]

#         processed_data = json.loads(generatePrompt(chunk_data)) 
#         temp.append(processed_data)  

#         all_chunks[f"chunk_{i + 1}"] = chunk_data

    
#     table_data = []


#     all_months = set()

#     # Gather all unique months across all chunks
#     for chunk in temp:
#         for activity, months in chunk.items():
#             all_months.update(months.keys())  # Get all unique month names

#     # Sort months in order
#     all_months = sorted(all_months, key=lambda x: pd.to_datetime(x, format='%b').month)

#     # Now build the rows for the DataFrame
#     for chunk in temp:
#         for activity, months in chunk.items():
#             row = {'Activity Name': activity}
#             total_count = 0  # Initialize total count for each activity
#             for month in all_months:
#                 count = months.get(month, 0)  # Default to 0 if the month is not present for the activity
#                 row[month] = count
#                 total_count += count  # Add the month count to the total
#             row['Total Count'] = total_count  # Add total count for the activity
#             table_data.append(row)

#     # Convert the table_data into a DataFrame
#     df = pd.DataFrame(table_data)

#     # Display the DataFrame in Streamlit
#     st.write("Activity Counts Table by Month")
#     st.dataframe(df)


    


# # File upload and processing
# uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

# if uploaded_file is not None and st.session_state.processed_df is None:
#     workbook = openpyxl.load_workbook(uploaded_file)
#     # sheet = workbook["TOWER 4 FINISHING."]
#     if "TOWER 4 FINISHING." in  workbook.sheetnames:
#         sheet = workbook["TOWER 4 FINISHING."]
#         activity_col_idx = 5  

#         non_bold_rows = []
#         for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16):
#             cell = row[activity_col_idx]  
#             if cell.font and not cell.font.b:  
#                 non_bold_rows.append(row_idx)

#         df = pd.read_excel(uploaded_file, sheet_name="TOWER 4 FINISHING.", skiprows=15)
        
#         df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
#                     'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
#                     'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
        
#         required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
#         df = df[required_columns]
        
#         df.index = df.index + 16
#         df = df.loc[df.index.isin(non_bold_rows)]

#         df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
#         df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)

#         df['Finish_Year'] = df['Finish'].dt.year
#         df['Finish_Month'] = df['Finish'].dt.month
#         df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')

#         st.session_state.processed_df = df
#     else:
#         activity_col_idx = 5 
#         sheet = workbook["M7 T5"]
#         non_bold_rows = []
        
#         for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16):
#             cell = row[activity_col_idx]  
#             if cell.font and not cell.font.b:  
#                 non_bold_rows.append(row_idx)
                
#         df = pd.read_excel(uploaded_file, sheet_name="M7 T5", skiprows=1)
        
#         expected_columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
#                    'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
#                    'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
        
#         actual_cols = len(df.columns)
# # print(f"Number of columns in Excel file: {actual_cols}")  # Debug info


#         if actual_cols >= len(expected_columns):
#             df.columns = expected_columns + [f'Extra_{i}' for i in range(actual_cols - len(expected_columns))]
#         else:
#             df.columns = expected_columns[:actual_cols]

#         required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
#         available_columns = [col for col in required_columns if col in df.columns]
#         df = df[available_columns]

#         df.index = df.index + 16
#         df = df.loc[df.index.isin(non_bold_rows)]


#         if 'Start' in df.columns:
#             df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
#         if 'Finish' in df.columns:
#             df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)
#             df['Finish_Year'] = df['Finish'].dt.year
#             df['Finish_Month'] = df['Finish'].dt.month
#             df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')

#         st.session_state.processed_df = df

# def to_excel(df):
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         df.to_excel(writer, index=True, sheet_name='Activity Counts')
#     excel_data = output.getvalue()
#     return excel_data

# # Process the filtered data
# if st.session_state.processed_df is not None:
#     df = st.session_state.processed_df
#     available_years = sorted(df['Finish_Year'].unique())
#     available_months = sorted(df['Finish_Month_Name'].unique())

#     selected_year = st.sidebar.selectbox('Select Year', available_years, index=available_years.index(df['Finish_Year'].max()))
#     selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

#     filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]

#     st.write(f"Filtered rows based on the selected months and year: {', '.join(selected_months)} {selected_year}")
#     st.write(filtered_df)
#     st.write(f"Number of rows: {len(filtered_df)}")

# #------------------------THIS IS JSON DATA---------------------

# #ithu ennoda json
#     result = filtered_df[['Activity Name', 'Finish_Month_Name']]

#     result_json = result.to_dict(orient='records')

#     # st.text(result_json)
#     df = pd.DataFrame(result_json)
    
#     mar_count = len(df[(df['Activity Name'] == 'EL-Third Fix ') & (df['Finish_Month_Name'] == 'Mar')])
    
#     st.write(f"Number of occurrences of 'Mar' for 'EL-Third Fix': {mar_count}")

#     month_order = [
#     'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
#     'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
# ]

# #------------------------THIS IS JSON DATA---------------------

#     activity_month_counts = pd.pivot_table(
#         filtered_df, 
#         values='Activity ID', 
#         index='Activity Name', 
#         columns='Finish_Month_Name', 
#         aggfunc='count', 
#         fill_value=0
#     )

#     existing_months = [month for month in month_order if month in activity_month_counts.columns]

#     activity_month_counts = activity_month_counts.reindex(columns=existing_months)

#     activity_month_counts['Total Count'] = activity_month_counts.sum(axis=1)

#     st.write(f"Activity counts by month for {selected_year}:")
#     st.write(activity_month_counts)
#     excel_file = to_excel(activity_month_counts)
#     st.download_button(
#     label="Download as Excel",
#     data=excel_file,
#     file_name=f"activity_counts_{selected_year}.xlsx",
#     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# )
    

# if st.button('Count The activity'):
#     createChunk(result_json)


import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO

st.title("Excel File Reader with Month and Year Filter")

WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

if 'processed_df' not in st.session_state:
    st.session_state.processed_df = None

if 'total_count_df' not in st.session_state:
    st.session_state.total_count_df = None

# Function to get access token
def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']

# Generate prompt for Watson API
def generatePrompt(json_datas):
    body = {
        "input": f"""
        Parse this JSON data, where each entry represents a task with 'Activity Name' and 'Finish_Month_Name' fields:
        {json_datas}
Count the number of occurrences of each unique 'Activity Name' for each 'Finish_Month_Name' in the months {', '.join(selected_months)} for the selected year. 
Return a JSON object where:
- Keys are unique 'Activity Name' values.
- Values are dictionaries with keys as 'Finish_Month_Name' from {', '.join(selected_months)} and values as the count of occurrences.
- Include all months from {', '.join(selected_months)} for every 'Activity Name', setting the count to 0 if there are no occurrences in that month.
Example output:
{{
    "Install Windows": {{"Mar": 2, "Apr": 1, "May": 0}},
    "Paint Walls": {{"Mar": 1, "Apr": 3, "May": 0}}
}}
Return only the JSON object, no code, no explanation, just the formatted JSON.
        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    return response.json()['results'][0]['generated_text'].strip()

# Function to create chunks and process data
def createChunk(result_json, chunk_size=2000):
    num_rows = len(result_json)
    num_chunks = (num_rows + chunk_size - 1) // chunk_size  
    all_chunks = {} 
    temp = []

    st.write(f"Filtered rows in JSON format (split into {chunk_size}-row chunks):")
 
    for i in range(num_chunks):
        start_idx = i * chunk_size
        end_idx = min((i + 1) * chunk_size, num_rows)
        chunk_data = result_json[start_idx:end_idx]

        processed_data = json.loads(generatePrompt(chunk_data)) 
        temp.append(processed_data)  

        all_chunks[f"chunk_{i + 1}"] = chunk_data

    
    table_data = []
    all_months = set()

    # Gather all unique months across all chunks
    for chunk in temp:
        for activity, months in chunk.items():
            all_months.update(months.keys())  # Get all unique month names

    # Sort months in order
    all_months = sorted(all_months, key=lambda x: pd.to_datetime(x, format='%b').month)

    # Now build the rows for the DataFrame
    for chunk in temp:
        for activity, months in chunk.items():
            row = {'Activity Name': activity}
            total_count = 0  # Initialize total count for each activity
            for month in all_months:
                count = months.get(month, 0)  # Default to 0 if the month is not present for the activity
                row[month] = count
                total_count += count  # Add the month count to the total
            row['Total Count'] = total_count  # Add total count for the activity
            table_data.append(row)

    # Convert the table_data into a DataFrame
    df = pd.DataFrame(table_data)

    # Display the DataFrame in Streamlit
    st.write("Activity Counts Table by Month")
    st.dataframe(df)

# File upload and processing
uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

if uploaded_file is not None and st.session_state.processed_df is None:
    workbook = openpyxl.load_workbook(uploaded_file)
    if "TOWER 4 FINISHING." in workbook.sheetnames:
        sheet = workbook["TOWER 4 FINISHING."]
        activity_col_idx = 5  

        non_bold_rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16):
            cell = row[activity_col_idx]  
            if cell.font and not cell.font.b:  
                non_bold_rows.append(row_idx)

        df = pd.read_excel(uploaded_file, sheet_name="TOWER 4 FINISHING.", skiprows=15)
        
        df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
                    'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                    'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
        
        required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
        df = df[required_columns]
        
        df.index = df.index + 16
        df = df.loc[df.index.isin(non_bold_rows)]

        df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
        df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)

        df['Finish_Year'] = df['Finish'].dt.year
        df['Finish_Month'] = df['Finish'].dt.month
        df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')

        st.session_state.processed_df = df
    else:
        activity_col_idx = 5 
        sheet = workbook["M7 T5"]
        non_bold_rows = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=17, max_col=16), start=16):
            cell = row[activity_col_idx]  
            if cell.font and not cell.font.b:  
                non_bold_rows.append(row_idx)
                
        df = pd.read_excel(uploaded_file, sheet_name="M7 T5", skiprows=1)
        
        expected_columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
                   'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                   'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
        
        actual_cols = len(df.columns)
        if actual_cols >= len(expected_columns):
            df.columns = expected_columns + [f'Extra_{i}' for i in range(actual_cols - len(expected_columns))]
        else:
            df.columns = expected_columns[:actual_cols]

        required_columns = ['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']
        available_columns = [col for col in required_columns if col in df.columns]
        df = df[available_columns]

        df.index = df.index + 16
        df = df.loc[df.index.isin(non_bold_rows)]

        if 'Start' in df.columns:
            df['Start'] = pd.to_datetime(df['Start'], errors='coerce', dayfirst=True)
        if 'Finish' in df.columns:
            df['Finish'] = pd.to_datetime(df['Finish'], errors='coerce', dayfirst=True)
            df['Finish_Year'] = df['Finish'].dt.year
            df['Finish_Month'] = df['Finish'].dt.month
            df['Finish_Month_Name'] = df['Finish'].dt.strftime('%b')

        st.session_state.processed_df = df

def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, sheet_name='Activity Counts')
    excel_data = output.getvalue()
    return excel_data

# Process the filtered data
if st.session_state.processed_df is not None:
    df = st.session_state.processed_df
    available_years = sorted(df['Finish_Year'].unique())
    available_months = sorted(df['Finish_Month_Name'].unique())

    selected_year = st.sidebar.selectbox('Select Year', available_years, index=available_years.index(df['Finish_Year'].max()))
    selected_months = st.sidebar.multiselect('Select Months', available_months, default=available_months)

    filtered_df = df[(df['Finish_Year'] == selected_year) & (df['Finish_Month_Name'].isin(selected_months))]

    st.write(f"Filtered rows based on the selected months and year: {', '.join(selected_months)} {selected_year}")
    st.write(filtered_df)
    st.write(f"Number of rows: {len(filtered_df)}")

    # JSON data preparation
    result = filtered_df[['Activity Name', 'Finish_Month_Name']]
    result_json = result.to_dict(orient='records')

    df = pd.DataFrame(result_json)
    
    mar_count = len(df[(df['Activity Name'] == 'EL-Third Fix ') & (df['Finish_Month_Name'] == 'Mar')])
    
    st.write(f"Number of occurrences of 'Mar' for 'EL-Third Fix': {mar_count}")

    month_order = [
    'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
    ]

    activity_month_counts = pd.pivot_table(
        filtered_df, 
        values='Activity ID', 
        index='Activity Name', 
        columns='Finish_Month_Name', 
        aggfunc='count', 
        fill_value=0
    )

    existing_months = [month for month in month_order if month in activity_month_counts.columns]
    activity_month_counts = activity_month_counts.reindex(columns=existing_months)
    activity_month_counts['Total Count'] = activity_month_counts.sum(axis=1)

    st.write(f"Activity counts by month for {selected_year}:")
    st.write(activity_month_counts)
    excel_file = to_excel(activity_month_counts)
    st.download_button(
    label="Download as Excel",
    data=excel_file,
    file_name=f"activity_counts_{selected_year}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Save to main_output.xlsx for compatibility with home.py
    activity_month_counts.to_excel("main_output.xlsx", index=True, sheet_name="Activity Counts")
    st.success("Output saved as 'main_output.xlsx' with sheet 'Activity Counts'!")

if st.button('Count The activity'):
    createChunk(result_json)